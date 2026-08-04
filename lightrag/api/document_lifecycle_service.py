from __future__ import annotations

import asyncio
import csv
import hashlib
import json
import mimetypes
import shutil
from collections.abc import AsyncIterator, Mapping
from contextlib import asynccontextmanager
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Awaitable, Callable, Literal, cast
from urllib.parse import urlsplit
from uuid import uuid4

from lightrag.artifact_runtime import (
    PipelineArtifactBinding,
    PipelineTerminalOutcome,
    assert_no_runtime_artifact_payload,
    canonicalize_pipeline_logical_filename,
    commit_pipeline_attempt_if_current,
)
from lightrag.api.artifact_materialization import (
    ArtifactMaterializationLease,
    ArtifactMaterializer,
    MaterializedDocumentTree,
)
from lightrag.api.artifact_lifecycle import (
    ARTIFACT_LIFECYCLE_MAX_PAGE_SIZE,
    ArtifactCleanupManifestRecord,
    artifact_cleanup_idempotency_key,
    normalize_artifact_target_uri,
)
from lightrag.api.commit_reconciliation import (
    MetadataCommitOutcome,
    MetadataCommitOutcomeUnknownError,
    MetadataCommitReconciliation,
    await_cancellation_safe_reconciliation,
)
from lightrag.api.config import ArtifactCleanupConfig
from lightrag.api.config_version_service import (
    active_parser_runtime_config_from_version,
)
from lightrag.api.enterprise_auth import enterprise_auth_enabled, get_current_principal
from lightrag.api.job_service import (
    assert_active_kb_generation,
    prepare_kb_job_payload,
)
from lightrag.api.kb_service import (
    KnowledgeBaseConflictError,
    KnowledgeBaseNotFoundError,
    KnowledgeBaseRecord,
    KnowledgeBaseService,
    utc_now_iso,
)
from lightrag.api.metadata_store import (
    ActiveDocumentBuildJobError,
    ActiveDocumentDeleteJobError,
    ActiveDocumentParseJobError,
    ActiveDocumentReplaceJobError,
    ArtifactRecord,
    DocumentMutationClaimResult,
    DocumentMutationCommitResult,
    DocumentAttemptOwnershipError,
    DocumentRecord,
    DocumentSnapshotConflictError,
    JobRecord,
    MetadataRecordNotFoundError,
    SQLiteMetadataStore,
    _new_document_attempt_token,
    document_mutation_manifest_group_id,
    document_source_generation_id,
)
from lightrag.api.object_storage import (
    DisabledObjectStorage,
    ObjectStorage,
    ObjectStorageError,
    ObjectStorageNotFoundError,
)
from lightrag.api.postgres_metadata_store import PostgresMetadataStore
from lightrag.constants import (
    DOCLING_RAW_DIR_SUFFIX,
    FULL_DOCS_FORMAT_LIGHTRAG,
    FULL_DOCS_FORMAT_PENDING_PARSE,
    LIBREOFFICE_RAW_DIR_SUFFIX,
    MINERU_RAW_DIR_SUFFIX,
    PARSED_DIR_NAME,
    PARSED_DIR_SUFFIX,
    PARSER_ENGINE_DOCLING,
    PARSER_ENGINE_LEGACY,
    PARSER_ENGINE_MINERU,
    PARSER_ENGINE_NATIVE,
    SUPPORTED_PARSER_ENGINES,
)
from lightrag.parser.routing import (
    normalize_parser_engine,
    parser_engine_supports_suffix,
    parser_suffix,
    resolve_file_parser_directives,
    sanitize_process_options,
    validate_process_options,
)
from lightrag.utils import compute_mdhash_id, generate_track_id, logger
from lightrag.utils_pipeline import (
    canonicalize_parser_hinted_basename,
)

SourceType = Literal["upload", "text", "url", "import", "scan"]
SOURCE_TYPES: tuple[SourceType, ...] = ("upload", "text", "url", "import", "scan")
MetadataStore = SQLiteMetadataStore | PostgresMetadataStore

_PREVIEW_SCHEMA_VERSION = 1
_PREVIEW_TEXT_MAX_BYTES = 256 * 1024
_PREVIEW_TABLE_MAX_ROWS = 50
_PREVIEW_TABLE_MAX_COLS = 50
_PREVIEW_TEXT_SUFFIXES = {
    ".txt",
    ".md",
    ".mdx",
    ".rtf",
    ".tex",
    ".html",
    ".htm",
    ".csv",
    ".json",
    ".xml",
    ".yaml",
    ".yml",
    ".log",
    ".conf",
    ".ini",
    ".properties",
    ".sql",
    ".bat",
    ".sh",
    ".c",
    ".h",
    ".cpp",
    ".hpp",
    ".py",
    ".java",
    ".js",
    ".ts",
    ".swift",
    ".go",
    ".rb",
    ".php",
    ".css",
    ".scss",
    ".less",
}
AgentProfileDirtyCallback = Callable[[str, str, str], Awaitable[None]]

# Sanitization rule: drop only path separators, control characters, and
# characters that are unsafe inside a filename on common filesystems
# (``<>:"|?*`` plus ASCII < 0x20). CJK / Latin-extended / accented letters
# stay intact so two PDFs whose names differ only in CJK characters don't
# both collapse to ``_.pdf`` and collide downstream in LightRAG's
# filename-based dedup.
_FILENAME_FORBIDDEN_CHARS = set('<>:"|?*\\/')


def _sanitize_filename_char(char: str) -> str:
    if not char:
        return "_"
    code = ord(char)
    if code < 0x20 or code == 0x7F:
        return "_"
    if char in _FILENAME_FORBIDDEN_CHARS:
        return "_"
    return char


_PARSEABLE_ENGINES = {
    PARSER_ENGINE_LEGACY,
    PARSER_ENGINE_NATIVE,
    PARSER_ENGINE_MINERU,
    PARSER_ENGINE_DOCLING,
}
_MARKDOWN_SUFFIXES = {".md", ".markdown"}
_IMAGE_SUFFIXES = {
    ".bmp",
    ".gif",
    ".jpeg",
    ".jpg",
    ".png",
    ".svg",
    ".tif",
    ".tiff",
    ".webp",
}
_ROOT_FILE_ARTIFACT_TYPES = {
    "content_list.json": "content_list",
    "middle.json": "middle_json",
    "middle_json.json": "middle_json",
    "model.json": "model_json",
    "model_json.json": "model_json",
    "layout.pdf": "layout_pdf",
}
_RAW_BUNDLE_MANIFEST = "_manifest.json"


class DocumentLifecycleError(RuntimeError):
    pass


class DocumentSourceChecksumError(DocumentLifecycleError):
    """Materialized source bytes do not match durable document metadata."""


class DocumentCowError(DocumentLifecycleError):
    """Base error for object-authoritative document copy-on-write execution."""


class DocumentCowManifestPreparationError(DocumentCowError):
    """Exact cleanup authority could not be prepared without ambiguity."""


class DocumentCowCommitOutcomeUnknownError(DocumentCowError):
    """A document COW commit could not be proved committed or rolled back."""

    def __init__(
        self,
        operation: Literal["replace", "delete"],
        *,
        document_id: str,
        job_id: str,
        attempt_token: str,
        manifest_group_id: str,
        reason: str | None = None,
    ) -> None:
        self.operation = operation
        self.document_id = document_id
        self.job_id = job_id
        self.attempt_token = attempt_token
        self.manifest_group_id = manifest_group_id
        self.reason = reason
        super().__init__(
            f"Document {operation} commit outcome is unknown; retry is fenced until "
            "exact reconciliation succeeds"
        )


class DocumentCowRetryableError(DocumentCowError):
    """A fenced COW attempt remains safe to re-drive."""

    def __init__(self, operation: str, *, document_id: str, attempt_token: str) -> None:
        self.operation = operation
        self.document_id = document_id
        self.attempt_token = attempt_token
        super().__init__(f"Document {operation} did not commit and is safe to retry")


class DocumentCowCompensationError(DocumentCowError):
    """Rollback was proved but candidate compensation could not be enqueued."""

    def __init__(self, *, document_id: str, job_id: str, attempt_token: str) -> None:
        self.document_id = document_id
        self.job_id = job_id
        self.attempt_token = attempt_token
        super().__init__(
            "Replacement rollback is safe, but candidate cleanup compensation "
            "requires orphan reconciliation"
        )


class DocumentCowEngineDeleteError(DocumentCowError):
    """The idempotent engine delete failed before metadata finalization."""

    def __init__(
        self,
        operation: Literal["replace", "delete"],
        *,
        document_id: str,
        job_id: str,
        attempt_token: str,
    ) -> None:
        self.operation = operation
        self.document_id = document_id
        self.job_id = job_id
        self.attempt_token = attempt_token
        self.result: DocumentReplaceCowResult | None = None
        super().__init__(
            f"Document {operation} engine cleanup did not complete; the attempt "
            "remains durably fenced"
        )


@dataclass(frozen=True, slots=True)
class DocumentReplaceCowResult:
    document: DocumentRecord
    attempt_token: str
    manifest_group_id: str
    manifest_ids: tuple[str, ...]
    cleanup_pending_count: int
    cleanup_retained_count: int
    cleanup_blocked_count: int
    source_generation_id: str
    phase: Literal["engine_cleanup_pending", "completed"]
    outcome: Literal["cleanup_pending", "completed"]


@dataclass(frozen=True, slots=True)
class DocumentDeleteCowResult:
    document: DocumentRecord
    attempt_token: str
    manifest_group_id: str
    manifest_ids: tuple[str, ...]
    cleanup_pending_count: int
    cleanup_retained_count: int
    cleanup_blocked_count: int
    phase: Literal["committed"] = "committed"
    outcome: Literal["deleted"] = "deleted"


@dataclass(frozen=True, slots=True)
class _DocumentCowTargetAuthority:
    target_uri: str
    target_kind: Literal["object", "prefix"]
    target_namespace: Literal["source", "legacy_source", "artifact"]
    artifact_id: str | None = None
    source_generation_id: str | None = None
    disposition: Literal["delete", "retain"] = "delete"


@dataclass(frozen=True, slots=True)
class _DocumentCowManifestTarget:
    """One exact cleanup target derived from a document mutation claim."""

    authority: _DocumentCowTargetAuthority
    expected_size_bytes: int | None = None
    expected_checksum: str | None = None
    expected_etag: str | None = None
    expected_version_id: str | None = None


# Operation prefix embedded in deterministic manifest ids so replace and
# delete manifests for the same document never collide.
_DOCUMENT_COW_OPERATION_PREFIX = {"replace": "r", "delete": "d"}


def _document_cow_source_object_key(
    workspace: str,
    document_id: str,
    source_generation_id: str,
    source_name: str,
) -> str:
    """Direct-final immutable candidate key for one replace attempt."""

    safe_name = "".join(_sanitize_filename_char(ch) for ch in source_name) or "source"
    return (
        f"workspaces/{workspace}/documents/{document_id}/source/generations/"
        f"{source_generation_id}/{safe_name}"
    )


def _document_cow_manifest_id(
    operation: Literal["replace", "delete"],
    *,
    manifest_group_id: str,
    idempotency_key: str,
) -> str:
    """Deterministic, unique manifest id for one COW cleanup target."""

    prefix = _DOCUMENT_COW_OPERATION_PREFIX.get(operation, "x")
    digest = hashlib.sha256(
        f"{manifest_group_id}|{idempotency_key}".encode("utf-8")
    ).hexdigest()
    return f"dcow-{prefix}-{digest[:24]}"


def _document_cow_engine_identity(
    *,
    kb_id: str,
    document_id: str,
    source_generation_id: str,
) -> str:
    """Deterministic engine-side identity for one object-mode mutation."""

    digest = hashlib.sha256(
        json.dumps(
            {
                "kb_id": kb_id,
                "document_id": document_id,
                "source_generation_id": source_generation_id,
            },
            separators=(",", ":"),
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    return f"engdoc-{digest[:24]}"


def _document_cow_orphan_reconcile_group_id(
    *,
    kb_id: str,
    kb_generation: str,
    document_id: str,
    job_id: str,
    attempt_token: str,
    candidate_uri: str,
) -> str:
    """Deterministic manifest group for one rolled-back candidate compensation."""

    digest = hashlib.sha256(
        json.dumps(
            {
                "kind": "orphan_reconcile",
                "kb_id": kb_id,
                "kb_generation": kb_generation,
                "document_id": document_id,
                "job_id": job_id,
                "attempt_token": attempt_token,
                "candidate_uri": normalize_artifact_target_uri(candidate_uri),
            },
            separators=(",", ":"),
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    return f"orphan-cow-{digest[:24]}"


def _extract_sha256_hex(source_hash: str) -> str | None:
    """Return a canonical hex SHA-256 for the immutable upload proof, or None.

    Accepts ``sha256:<hex>`` (the durable document hash format) or a bare hex
    digest.  Non-canonical values return None so the proof is skipped rather
    than rejected, matching Store A's permissive checksum handling.
    """

    if not isinstance(source_hash, str) or not source_hash:
        return None
    candidate = source_hash.strip()
    if candidate.lower().startswith("sha256:"):
        candidate = candidate.split(":", 1)[1]
    elif candidate.lower().startswith("sha-256:"):
        candidate = candidate.split(":", 1)[1]
    if len(candidate) == 64 and all(ch in "0123456789abcdefABCDEF" for ch in candidate):
        return candidate.lower()
    return None


def _document_delete_cow_result_from_commit(
    committed: DocumentMutationCommitResult,
    attempt_token: str,
) -> DocumentDeleteCowResult:
    return DocumentDeleteCowResult(
        document=committed.document,
        attempt_token=attempt_token,
        manifest_group_id=committed.manifest_group_id,
        manifest_ids=committed.manifest_ids,
        cleanup_pending_count=committed.pending_cleanup_count,
        cleanup_retained_count=committed.retained_cleanup_count,
        cleanup_blocked_count=committed.blocked_cleanup_count,
    )


def _resolve_service_attempt_owner(
    document: DocumentRecord,
    *,
    operation: Literal["parse", "build"],
    job_id: str,
    claim_token: str | None,
    strict: bool,
) -> tuple[str | None, Literal["pending", "current"] | None]:
    """Resolve a persisted owner for local compatibility without bypassing CAS."""

    pending_status = f"{operation}_queued"
    current_status = "parsing" if operation == "parse" else "building"
    if document.status == pending_status:
        phase: Literal["pending", "current"] | None = "pending"
    elif document.status == current_status:
        phase = "current"
    else:
        if strict and claim_token is not None:
            raise DocumentAttemptOwnershipError(
                f"document_{operation}_attempt",
                document.id,
                expected={
                    "status": current_status,
                    "job_id": job_id,
                    "claim_token": claim_token,
                },
                current={
                    "status": document.status,
                    "job_id": None,
                    "claim_token": None,
                },
            )
        return claim_token, None

    owner_job_value = document.metadata.get(f"{phase}_{operation}_job_id")
    owner_token_value = document.metadata.get(f"{phase}_{operation}_claim_token")
    owner_job_id = (
        owner_job_value
        if isinstance(owner_job_value, str) and owner_job_value
        else None
    )
    owner_claim_token = (
        owner_token_value
        if isinstance(owner_token_value, str) and owner_token_value
        else None
    )

    if claim_token is not None:
        if strict and (owner_job_id, owner_claim_token) != (job_id, claim_token):
            raise DocumentAttemptOwnershipError(
                f"document_{operation}_attempt",
                document.id,
                expected={
                    "status": pending_status if phase == "pending" else current_status,
                    "job_id": job_id,
                    "claim_token": claim_token,
                },
                current={
                    "status": document.status,
                    "job_id": owner_job_id,
                    "claim_token": owner_claim_token,
                },
            )
        return claim_token, phase

    if owner_claim_token is not None and owner_job_id == job_id:
        return owner_claim_token, phase
    if strict and owner_job_id not in {None, job_id}:
        raise DocumentAttemptOwnershipError(
            f"document_{operation}_attempt",
            document.id,
            expected={
                "status": pending_status if phase == "pending" else current_status,
                "job_id": job_id,
                "claim_token": None,
            },
            current={
                "status": document.status,
                "job_id": owner_job_id,
                "claim_token": owner_claim_token,
            },
        )
    return None, phase


def _active_document_job_error(
    document: DocumentRecord,
) -> (
    ActiveDocumentParseJobError
    | ActiveDocumentBuildJobError
    | ActiveDocumentDeleteJobError
    | ActiveDocumentReplaceJobError
    | None
):
    if document.status == "parse_queued":
        job_id = document.metadata.get("pending_parse_job_id")
        return ActiveDocumentParseJobError(document.id, str(job_id or "unknown"))
    if document.status == "parsing":
        job_id = document.metadata.get("current_parse_job_id")
        return ActiveDocumentParseJobError(document.id, str(job_id or "unknown"))
    if document.status == "build_queued":
        job_id = document.metadata.get("pending_build_job_id")
        return ActiveDocumentBuildJobError(document.id, str(job_id or "unknown"))
    if document.status == "building":
        job_id = document.metadata.get("current_build_job_id")
        return ActiveDocumentBuildJobError(document.id, str(job_id or "unknown"))
    if document.status == "deleting":
        job_id = document.metadata.get(
            "pending_delete_job_id"
        ) or document.metadata.get("current_delete_job_id")
        return ActiveDocumentDeleteJobError(document.id, str(job_id or "unknown"))
    if document.status == "replacing":
        job_id = document.metadata.get(
            "pending_replace_job_id"
        ) or document.metadata.get("current_replace_job_id")
        return ActiveDocumentReplaceJobError(document.id, str(job_id or "unknown"))
    return None


def _active_document_job_error_code(
    exc: ActiveDocumentParseJobError
    | ActiveDocumentBuildJobError
    | ActiveDocumentDeleteJobError
    | ActiveDocumentReplaceJobError,
) -> str:
    if isinstance(exc, ActiveDocumentParseJobError):
        return "parse_job_active"
    if isinstance(exc, ActiveDocumentBuildJobError):
        return "build_job_active"
    if isinstance(exc, ActiveDocumentDeleteJobError):
        return "delete_job_active"
    return "replace_job_active"


@dataclass(slots=True)
class DocumentSourceInput:
    source_name: str
    content: bytes
    source_type: SourceType
    content_type: str | None = None
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class DocumentBatchResult:
    job: JobRecord
    batch_id: str
    documents: list[DocumentRecord]
    created: bool = True


@dataclass(slots=True)
class DocumentParsePlan:
    document: DocumentRecord
    source_name: str
    source_object_uri: str | None
    raw_object_refs: tuple["DocumentRawObjectReference", ...]
    parser_engine: str
    process_options: str
    parser_hash: str
    lightrag_doc_id: str
    kb_generation: str = ""
    job_id: str | None = None
    force_reparse: bool = False
    auto_index: bool = False
    expected_status: str | None = None
    expected_source_hash: str | None = None
    expected_parser_hash: str | None = None
    expected_current_parse_generation_id: str | None = None
    expected_current_sidecar_artifact_id: str | None = None
    expected_current_blocks_artifact_id: str | None = None
    expected_index_hash: str | None = None
    claim_token: str | None = None
    execution: "DocumentParseExecution | None" = field(
        default=None, repr=False, compare=False
    )

    def __post_init__(self) -> None:
        if self.expected_status is None:
            self.expected_status = self.document.status
            self.expected_source_hash = self.document.source_hash
            self.expected_parser_hash = self.document.parser_hash
            self.expected_current_parse_generation_id = self.document.metadata.get(
                "current_parse_generation_id"
            )
            self.expected_current_sidecar_artifact_id = self.document.metadata.get(
                "current_sidecar_artifact_id"
            )
            self.expected_current_blocks_artifact_id = self.document.metadata.get(
                "current_blocks_artifact_id"
            )
            self.expected_index_hash = self.document.index_hash
        if self.claim_token is None:
            token_key = (
                "pending_parse_claim_token"
                if self.document.status == "parse_queued"
                else "current_parse_claim_token"
                if self.document.status == "parsing"
                else None
            )
            token = self.document.metadata.get(token_key) if token_key else None
            if isinstance(token, str) and token:
                self.claim_token = token

    @property
    def expected_snapshot(self) -> dict[str, Any]:
        return {
            "status": self.expected_status,
            "source_hash": self.expected_source_hash,
            "parser_hash": self.expected_parser_hash,
            "current_parse_generation_id": (self.expected_current_parse_generation_id),
            "current_sidecar_artifact_id": (self.expected_current_sidecar_artifact_id),
            "current_blocks_artifact_id": (self.expected_current_blocks_artifact_id),
            "index_hash": self.expected_index_hash,
        }


@dataclass(frozen=True, slots=True)
class DocumentRawObjectReference:
    artifact_id: str
    object_prefix_uri: str
    directory_name: str
    checksum: str | None


@dataclass(slots=True)
class DocumentParseExecution:
    lease: ArtifactMaterializationLease | None
    scratch_document_root: Path
    source_path: Path
    parsed_tree: Path
    canonical_document_root: Path
    producer_task: asyncio.Task[dict[str, Any]] | None = field(
        default=None, repr=False, compare=False
    )

    def canonical_path_for(self, runtime_path: Path | str) -> Path:
        runtime = Path(runtime_path).resolve(strict=False)
        scratch_root = self.scratch_document_root.resolve(strict=False)
        try:
            relative = runtime.relative_to(scratch_root)
        except ValueError as exc:
            raise DocumentLifecycleError(
                "Parser artifact path escapes the materialized document tree"
            ) from exc
        canonical = (self.canonical_document_root / relative).resolve(strict=False)
        canonical_root = self.canonical_document_root.resolve(strict=False)
        if not canonical.is_relative_to(canonical_root):
            raise DocumentLifecycleError(
                "Canonical parser artifact locator escapes the document root"
            )
        return canonical

    def defer_cleanup(self) -> None:
        if self.lease is not None and not self.lease.cleanup_deferred:
            self.lease.defer_cleanup()

    def cleanup(self) -> None:
        if self.lease is not None and not self.lease.cleanup_deferred:
            self.lease.cleanup()

    def durable_error_message(self, error: object) -> str:
        message = str(error)
        if self.lease is None:
            return message
        replacements = (
            (
                self.scratch_document_root.resolve(strict=False).as_uri(),
                self.canonical_document_root.resolve(strict=False).as_uri(),
            ),
            (str(self.scratch_document_root), str(self.canonical_document_root)),
            (
                self.lease.path.resolve(strict=False).as_uri(),
                "materialization://redacted",
            ),
            (str(self.lease.path), "<artifact-materialization>"),
        )
        for runtime_value, durable_value in replacements:
            message = message.replace(runtime_value, durable_value)
        return message.replace(".lightrag-scratch", "artifact-materialization")


@dataclass(slots=True)
class DocumentBatchParsePlan:
    batch_id: str
    plans: list[DocumentParsePlan]
    failures: list[dict[str, Any]]


@dataclass(slots=True)
class DocumentParseResult:
    document: DocumentRecord
    artifacts: list[ArtifactRecord]


@dataclass(slots=True)
class PendingArtifact:
    record: ArtifactRecord
    runtime_path: Path
    is_directory: bool


@dataclass(frozen=True, slots=True)
class UploadedArtifactObject:
    uri: str
    is_prefix: bool


@dataclass(slots=True)
class DocumentDeleteFileResult:
    deleted_source: bool = False
    deleted_artifacts: list[str] = field(default_factory=list)
    deleted_objects: list[str] = field(default_factory=list)
    skipped: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


@dataclass(slots=True)
class DocumentReplacementSource:
    source_name: str
    content: bytes
    source_type: SourceType
    source_hash: str
    content_type: str | None
    size_bytes: int


@dataclass(slots=True)
class ArtifactFileResult:
    artifact: ArtifactRecord
    path: Path
    filename: str
    media_type: str
    is_directory: bool = False
    lease: ArtifactMaterializationLease | None = None

    def cleanup(self) -> None:
        if self.lease is not None and not self.lease.cleanup_deferred:
            self.lease.cleanup()

    def defer_cleanup(self) -> None:
        if self.lease is not None and not self.lease.cleanup_deferred:
            self.lease.defer_cleanup()


@dataclass(slots=True)
class ArtifactDownloadUrlResult:
    artifact: ArtifactRecord
    url: str
    object_uri: str
    expires_in_seconds: int
    filename: str
    media_type: str


class DocumentLifecycleService:
    def __init__(
        self,
        kb_service: KnowledgeBaseService,
        metadata_store: MetadataStore,
        source_root: str | Path,
        object_storage: ObjectStorage | None = None,
        agent_profile_dirty_callback: AgentProfileDirtyCallback | None = None,
        *,
        artifact_storage_mode: str = "local",
        materializer: ArtifactMaterializer | None = None,
        artifact_cleanup_config: ArtifactCleanupConfig | None = None,
        clock: Callable[[], datetime] | None = None,
    ):
        normalized_storage_mode = str(artifact_storage_mode or "local").strip().lower()
        if normalized_storage_mode not in {"local", "object"}:
            raise DocumentLifecycleError(
                "artifact_storage_mode must be either 'local' or 'object'"
            )
        if normalized_storage_mode == "object":
            if object_storage is None or isinstance(
                object_storage, DisabledObjectStorage
            ):
                raise DocumentLifecycleError(
                    "Object artifact mode requires remote object storage"
                )
            if not isinstance(materializer, ArtifactMaterializer):
                raise DocumentLifecycleError(
                    "Object artifact mode requires an explicit validated "
                    "ArtifactMaterializer"
                )
            if materializer.object_storage is not object_storage:
                raise DocumentLifecycleError(
                    "Document lifecycle and materializer must share object storage"
                )
            requested_root = Path(source_root).expanduser().resolve(strict=False)
            if materializer.input_root != requested_root:
                raise DocumentLifecycleError(
                    "Document lifecycle source root conflicts with materializer INPUT_DIR"
                )
        self._kb_service = kb_service
        self._metadata_store = metadata_store
        self._source_root = Path(source_root)
        self._object_storage = object_storage
        self._artifact_storage_mode = normalized_storage_mode
        self._materializer = materializer
        self._agent_profile_dirty_callback = agent_profile_dirty_callback
        if artifact_cleanup_config is not None and not isinstance(
            artifact_cleanup_config, ArtifactCleanupConfig
        ):
            raise DocumentLifecycleError(
                "artifact_cleanup_config must be an ArtifactCleanupConfig"
            )
        self._artifact_cleanup_config = (
            artifact_cleanup_config or ArtifactCleanupConfig()
        )
        self._clock = clock or (lambda: datetime.now(timezone.utc))

    def set_agent_profile_dirty_callback(
        self, callback: AgentProfileDirtyCallback | None
    ) -> None:
        self._agent_profile_dirty_callback = callback

    @property
    def metadata_store(self) -> MetadataStore:
        return self._metadata_store

    @property
    def kb_service(self) -> KnowledgeBaseService:
        return self._kb_service

    @property
    def source_root(self) -> Path:
        return self._source_root

    @property
    def artifact_storage_mode(self) -> str:
        return self._artifact_storage_mode

    @property
    def object_authoritative(self) -> bool:
        return self._artifact_storage_mode == "object"

    @property
    def object_storage(self) -> ObjectStorage | None:
        return self._object_storage

    @property
    def materializer(self) -> ArtifactMaterializer | None:
        return self._materializer

    def canonical_document_root(self, document: DocumentRecord) -> Path:
        return self._canonical_document_root(document)

    async def release_pipeline_artifact_attempt_if_owned(
        self,
        binding: PipelineArtifactBinding,
        outcome: PipelineTerminalOutcome,
    ) -> DocumentRecord | None:
        """Release only the still-current parse/build attempt from a session.

        The generation fence prevents an old runtime session from touching a
        recreated KB.  The metadata-store release operations are themselves
        owner-aware, so a newer job/token winner is returned unchanged.
        """

        if outcome not in {
            PipelineTerminalOutcome.FAILED,
            PipelineTerminalOutcome.CANCELLED,
        }:
            raise DocumentLifecycleError(
                "Pipeline artifact release requires failed or cancelled outcome"
            )
        error_code = (
            "pipeline_artifact_cancelled"
            if outcome is PipelineTerminalOutcome.CANCELLED
            else "pipeline_artifact_failed"
        )
        error_message = (
            "Pipeline artifact processing was cancelled"
            if outcome is PipelineTerminalOutcome.CANCELLED
            else "Pipeline artifact processing failed"
        )
        try:
            async with self.kb_write_guard(
                binding.kb_id,
                expected_generation=binding.kb_generation,
            ) as record:
                metadata_patch = {
                    f"last_failed_{binding.operation}_job_id": binding.job_id,
                }
                if binding.operation == "parse":
                    return await self._metadata_store.release_document_parse_if_owned(
                        record.id,
                        binding.document_id,
                        job_id=binding.job_id,
                        claim_token=binding.claim_token,
                        error_code=error_code,
                        error_message=error_message,
                        metadata_patch=metadata_patch,
                    )
                return await self._metadata_store.release_document_build_if_owned(
                    record.id,
                    binding.document_id,
                    job_id=binding.job_id,
                    claim_token=binding.claim_token,
                    error_code=error_code,
                    error_message=error_message,
                    metadata_patch=metadata_patch,
                )
        except (
            KnowledgeBaseConflictError,
            KnowledgeBaseNotFoundError,
            MetadataRecordNotFoundError,
        ):
            # A replaced/deleted generation or document is already a stale
            # loser.  Never let its late cleanup overwrite the current winner.
            return None

    async def compensate_uploaded_artifact_objects(
        self, uploaded: list[UploadedArtifactObject]
    ) -> None:
        await self._compensate_uploaded_artifact_objects(uploaded)

    def assert_destructive_operation_supported(self, operation: str) -> None:
        if self.object_authoritative:
            raise DocumentLifecycleError(
                f"{operation} is disabled in object artifact mode until Phase 3"
            )

    async def _active_parser_defaults_for_record(self, record: Any) -> dict[str, str]:
        if not record.active_config_version_id:
            return {}
        active_config_version = await self._metadata_store.get_config_version(
            record.id, record.active_config_version_id
        )
        return active_parser_runtime_config_from_version(active_config_version)

    @asynccontextmanager
    async def kb_write_guard(
        self,
        kb_id: str,
        *,
        expected_generation: str | None = None,
    ) -> AsyncIterator[KnowledgeBaseRecord]:
        """Hold the shared deletion fence for one direct service mutation."""

        captured = await self._kb_service.get(kb_id, include_deleted=True)
        generation = expected_generation or captured.generation
        if captured.generation != generation:
            raise KnowledgeBaseConflictError(
                f"Knowledge base '{captured.id}' changed generation"
            )
        async with self._metadata_store.kb_write_guard(captured.id, generation):
            current = await assert_active_kb_generation(
                self._kb_service,
                self._metadata_store,
                captured.id,
                generation,
            )
            if current.workspace != captured.workspace:
                raise KnowledgeBaseConflictError(
                    f"Knowledge base '{current.id}' changed workspace"
                )
            yield current

    async def create_source_batch(
        self,
        kb_id: str,
        sources: list[DocumentSourceInput],
        *,
        auto_parse: bool = False,
        auto_index: bool = False,
        parser_engine: str | None = None,
        process_options: str | None = None,
        idempotency_key: str | None = None,
    ) -> DocumentBatchResult:
        if not sources:
            raise ValueError("At least one document source is required")

        job_type = "parse" if auto_parse else "upload"
        record, generation_payload = await prepare_kb_job_payload(
            self._kb_service,
            self._metadata_store,
            kb_id,
            job_type=job_type,
            payload=None,
        )
        async with self.kb_write_guard(
            record.id,
            expected_generation=record.generation,
        ) as guarded_record:
            if guarded_record.workspace != record.workspace:
                raise KnowledgeBaseConflictError(
                    f"Knowledge base '{record.id}' changed workspace"
                )
            return await self._create_source_batch_guarded(
                record,
                generation_payload,
                sources,
                auto_parse=auto_parse,
                auto_index=auto_index,
                parser_engine=parser_engine,
                process_options=process_options,
                idempotency_key=idempotency_key,
            )

    async def _create_source_batch_guarded(
        self,
        record: KnowledgeBaseRecord,
        generation_payload: dict[str, Any],
        sources: list[DocumentSourceInput],
        *,
        auto_parse: bool = False,
        auto_index: bool = False,
        parser_engine: str | None = None,
        process_options: str | None = None,
        idempotency_key: str | None = None,
    ) -> DocumentBatchResult:
        if not sources:
            raise ValueError("At least one document source is required")

        job_type = "parse" if auto_parse else "upload"
        if auto_parse:
            defaults = await self._active_parser_defaults_for_record(record)
            parser_engine, process_options = _apply_parse_defaults(
                parser_engine,
                process_options,
                defaults,
            )
        workspace_dir = self._source_root / record.workspace
        workspace_dir.mkdir(parents=True, exist_ok=True)
        batch_id = generate_track_id("batch")
        job_id = generate_track_id(f"job_{job_type}")
        document_status = "parse_queued" if auto_parse else "uploaded"
        now = utc_now_iso()
        # Stamp the uploading principal (enterprise mode only) so editors can
        # self-delete their own uploads; reads the request-scoped contextvar set
        # by combined_auth, mirroring JobService's principal attribution.
        created_by: str | None = None
        if enterprise_auth_enabled():
            principal = get_current_principal()
            if principal is not None:
                created_by = principal.user_id
        saved_paths: list[Path] = []
        saved_dirs: list[Path] = []
        saved_object_uris: list[str] = []
        documents: list[DocumentRecord] = []
        source_fingerprints: list[dict[str, Any]] = []
        metadata_commit_uncertain = False

        try:
            for source in sources:
                if not source.content:
                    raise ValueError("Document content cannot be empty")
                safe_name = _sanitize_source_name(source.source_name)
                content_hash = _content_hash(source.content)
                source_fingerprints.append(
                    {
                        "source_name": safe_name,
                        "source_type": source.source_type,
                        "content_type": source.content_type,
                        "source_hash": content_hash,
                        "metadata": source.metadata,
                    }
                )
                document_id = f"doc_{uuid4().hex[:12]}"
                target_path = _write_source_file(
                    workspace_dir, document_id, safe_name, source.content
                )
                source_object_uri = await self._persist_source_file(
                    record.workspace,
                    document_id,
                    target_path,
                    content_type=source.content_type,
                )
                if self.object_authoritative and not source_object_uri:
                    raise DocumentLifecycleError(
                        "Object artifact mode requires a durable source object URI"
                    )
                if source_object_uri:
                    saved_object_uris.append(source_object_uri)
                saved_paths.append(target_path)
                saved_dirs.append(target_path.parent)
                documents.append(
                    DocumentRecord(
                        id=document_id,
                        kb_id=record.id,
                        workspace=record.workspace,
                        lightrag_doc_id=None,
                        source_type=source.source_type,
                        source_name=safe_name,
                        source_uri=str(target_path),
                        source_hash=content_hash,
                        content_type=source.content_type,
                        size_bytes=len(source.content),
                        parser_hash=None,
                        index_hash=None,
                        status=document_status,
                        enabled=True,
                        archived=False,
                        chunks_count=None,
                        entity_count=None,
                        relation_count=None,
                        error_code=None,
                        error_message=None,
                        metadata={
                            **source.metadata,
                            **(
                                {"source_object_uri": source_object_uri}
                                if source_object_uri
                                else {}
                            ),
                            **({"created_by": created_by} if created_by else {}),
                            "batch_id": batch_id,
                            "auto_parse": auto_parse,
                            "auto_index": auto_index,
                            "parser_engine": parser_engine,
                            "process_options": process_options,
                            **(
                                {
                                    "pending_parse_job_id": job_id,
                                    "pending_parse_batch_id": batch_id,
                                    "pending_parse_claim_token": (
                                        _new_document_attempt_token("parse")
                                    ),
                                }
                                if auto_parse
                                else {}
                            ),
                        },
                        created_at=now,
                        updated_at=now,
                        deleted_at=None,
                    )
                )

            job = JobRecord(
                id=job_id,
                kb_id=record.id,
                workspace=record.workspace,
                batch_id=batch_id,
                document_id=None,
                job_type=job_type,
                status="queued" if auto_parse else "succeeded",
                stage="parsing" if auto_parse else "uploading",
                progress=0.0 if auto_parse else 1.0,
                total_items=len(documents),
                completed_items=0 if auto_parse else len(documents),
                failed_items=0,
                idempotency_key=idempotency_key,
                config_version_id=record.active_config_version_id,
                config_hash=None,
                retry_count=0,
                max_retries=3,
                payload={
                    **generation_payload,
                    "auto_parse": auto_parse,
                    "auto_index": auto_index,
                    "parser_engine": parser_engine,
                    "process_options": process_options,
                    "source_types": sorted({source.source_type for source in sources}),
                    "document_ids": [document.id for document in documents],
                    "idempotency_fingerprint": _idempotency_fingerprint(
                        {
                            "auto_parse": auto_parse,
                            "auto_index": auto_index,
                            "parser_engine": parser_engine,
                            "process_options": process_options,
                            "sources": source_fingerprints,
                        }
                    ),
                },
                result={"documents_created": len(documents)}
                if not auto_parse
                else None,
                error_code=None,
                error_message=None,
                created_at=now,
                updated_at=now,
                queued_at=now if auto_parse else None,
                started_at=now if not auto_parse else None,
                finished_at=now if not auto_parse else None,
                cancelled_at=None,
            )
            # Re-assert immediately before the atomic metadata write. HTTP
            # callers also hold the pure-ASGI shared fence across all file /
            # object staging; this second check protects direct service users
            # from persisting a job after a lifecycle transition.
            final_record = await assert_active_kb_generation(
                self._kb_service,
                self._metadata_store,
                record.id,
                record.generation,
            )
            if final_record.workspace != record.workspace:
                raise KnowledgeBaseConflictError(
                    f"Knowledge base '{record.id}' changed workspace"
                )
            metadata_commit_uncertain = True
            try:
                (
                    created_documents,
                    created_job,
                    created,
                ) = await self._metadata_store.create_documents_and_job(documents, job)
            except (Exception, asyncio.CancelledError) as commit_error:
                return await self._reconcile_initial_source_commit_exception(
                    documents=documents,
                    job=job,
                    saved_paths=saved_paths,
                    saved_dirs=saved_dirs,
                    saved_object_uris=saved_object_uris,
                    commit_error=commit_error,
                )
            metadata_commit_uncertain = False
            if not created:
                self._cleanup_saved_sources(saved_paths, saved_dirs)
                await self._cleanup_saved_objects(saved_object_uris)
            elif self.object_authoritative:
                # The object upload and atomic document/job commit are both
                # durable at this point.  Only now may stateless mode discard
                # the request-local source cache.
                self._cleanup_saved_sources(saved_paths, saved_dirs)
            return DocumentBatchResult(
                job=created_job,
                batch_id=created_job.batch_id or batch_id,
                documents=created_documents,
                created=created,
            )
        except BaseException:
            if not metadata_commit_uncertain:
                self._cleanup_saved_sources(saved_paths, saved_dirs)
                await self._cleanup_saved_objects(saved_object_uris)
            raise

    async def _read_initial_source_commit_outcome(
        self,
        *,
        documents: list[DocumentRecord],
        job: JobRecord,
    ) -> MetadataCommitReconciliation[DocumentBatchResult]:
        candidate_ids = [document.id for document in documents]
        (
            persisted_documents,
            persisted_job,
        ) = await self._metadata_store.get_documents_and_job_by_ids(
            job.kb_id,
            candidate_ids,
            job.id,
        )
        if not persisted_documents and persisted_job is None:
            return MetadataCommitReconciliation(
                outcome=MetadataCommitOutcome.ROLLED_BACK,
                reason="candidate_records_absent",
            )
        if len(persisted_documents) != len(documents) or persisted_job is None:
            return MetadataCommitReconciliation(
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="partial_candidate_presence",
            )

        persisted_by_id = {document.id: document for document in persisted_documents}
        if set(persisted_by_id) != set(candidate_ids):
            return MetadataCommitReconciliation(
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="candidate_document_identity_mismatch",
            )
        if not all(
            _initial_source_document_commit_matches(
                candidate,
                persisted_by_id[candidate.id],
            )
            for candidate in documents
        ) or not _initial_source_job_commit_matches(job, persisted_job):
            return MetadataCommitReconciliation(
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="candidate_record_mismatch",
            )

        ordered_documents = [
            persisted_by_id[document_id] for document_id in candidate_ids
        ]
        return MetadataCommitReconciliation(
            outcome=MetadataCommitOutcome.COMMITTED,
            value=DocumentBatchResult(
                job=persisted_job,
                batch_id=persisted_job.batch_id or job.batch_id or "",
                documents=ordered_documents,
                created=True,
            ),
            reason="candidate_records_match",
        )

    async def _reconcile_initial_source_commit_exception(
        self,
        *,
        documents: list[DocumentRecord],
        job: JobRecord,
        saved_paths: list[Path],
        saved_dirs: list[Path],
        saved_object_uris: list[str],
        commit_error: BaseException,
    ) -> DocumentBatchResult:
        async def read_back() -> MetadataCommitReconciliation[DocumentBatchResult]:
            return await self._read_initial_source_commit_outcome(
                documents=documents,
                job=job,
            )

        caller_cancelled = isinstance(commit_error, asyncio.CancelledError)
        readback_error: BaseException | None = None
        try:
            safe_result = await await_cancellation_safe_reconciliation(read_back)
            reconciliation = safe_result.value
            caller_cancelled = caller_cancelled or safe_result.caller_cancelled
        except asyncio.CancelledError as exc:
            readback_error = exc
            reconciliation = MetadataCommitReconciliation[DocumentBatchResult](
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="readback_cancelled",
            )
        except Exception as exc:  # noqa: BLE001 - read failure means unknown
            readback_error = exc
            reconciliation = MetadataCommitReconciliation[DocumentBatchResult](
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="readback_failed",
            )

        if (
            reconciliation.outcome is MetadataCommitOutcome.COMMITTED
            and reconciliation.value is not None
        ):
            if self.object_authoritative:
                self._cleanup_saved_sources(saved_paths, saved_dirs)
            if caller_cancelled:
                if isinstance(commit_error, asyncio.CancelledError):
                    raise commit_error
                raise asyncio.CancelledError() from commit_error
            return reconciliation.value

        if reconciliation.outcome is MetadataCommitOutcome.ROLLED_BACK:
            self._cleanup_saved_sources(saved_paths, saved_dirs)
            await self._cleanup_saved_objects(saved_object_uris)
            if caller_cancelled:
                if isinstance(commit_error, asyncio.CancelledError):
                    raise commit_error
                raise asyncio.CancelledError() from commit_error
            raise commit_error

        candidate_ids = [document.id for document in documents]
        logger.warning(
            "metadata_commit_reconciliation outcome=unknown operation=%s "
            "kb_id=%s candidate_job_id=%s candidate_document_ids=%s reason=%s "
            "commit_error_type=%s readback_error_type=%s",
            "initial_source_document_job",
            job.kb_id,
            job.id,
            candidate_ids,
            reconciliation.reason or "unknown",
            type(commit_error).__name__,
            type(readback_error).__name__ if readback_error is not None else None,
        )
        unknown_error = MetadataCommitOutcomeUnknownError(
            "initial_source_document_job",
            candidate_document_ids=candidate_ids,
            candidate_job_id=job.id,
            reason=reconciliation.reason,
        )
        if caller_cancelled:
            if isinstance(commit_error, asyncio.CancelledError):
                raise commit_error from unknown_error
            raise asyncio.CancelledError() from unknown_error
        raise unknown_error from (readback_error or commit_error)

    async def list_documents(
        self,
        kb_id: str,
        *,
        status: str | None = None,
        source_name: str | None = None,
        limit: int = 50,
        offset: int = 0,
    ) -> tuple[list[DocumentRecord], int]:
        record = await self._kb_service.get(kb_id)
        return await self._metadata_store.list_documents(
            record.id,
            status=status,
            source_name=source_name,
            limit=limit,
            offset=offset,
        )

    async def get_document(self, kb_id: str, document_id: str) -> DocumentRecord:
        record = await self._kb_service.get(kb_id)
        return await self._metadata_store.get_document(record.id, document_id)

    async def get_documents_by_ids(
        self, kb_id: str, document_ids: list[str]
    ) -> list[DocumentRecord]:
        record = await self._kb_service.get(kb_id)
        return await self._metadata_store.get_documents_by_ids(record.id, document_ids)

    async def get_documents_by_source_keys(
        self, kb_id: str, source_keys: list[str]
    ) -> dict[str, DocumentRecord]:
        record = await self._kb_service.get(kb_id)
        return await self._metadata_store.get_documents_by_source_keys(
            record.id, source_keys
        )

    async def update_document(
        self,
        kb_id: str,
        document_id: str,
        *,
        metadata_patch: dict[str, Any] | None = None,
        enabled: bool | None = None,
        archived: bool | None = None,
        expected_generation: str | None = None,
    ) -> DocumentRecord:
        async with self.kb_write_guard(
            kb_id, expected_generation=expected_generation
        ) as record:
            document = await self._metadata_store.update_document(
                record.id,
                document_id,
                metadata_patch=metadata_patch,
                enabled=enabled,
                archived=archived,
            )
            if enabled is not None or archived is not None:
                await self._notify_agent_profile_dirty(
                    record.id, document_id, "document_lifecycle_changed"
                )
            return document

    async def claim_delete(
        self,
        kb_id: str,
        document_id: str,
        *,
        job: JobRecord,
        delete_source_file: bool = False,
        delete_artifacts: bool = False,
    ) -> DocumentRecord:
        self.assert_destructive_operation_supported("Document delete")
        async with self.kb_write_guard(kb_id) as record:
            return await self._metadata_store.claim_document_deleting(
                record.id,
                document_id,
                metadata_patch={
                    "pending_delete_job_id": job.id,
                    "delete_source_file": delete_source_file,
                    "delete_artifacts": delete_artifacts,
                },
            )

    async def claim_batch_delete(
        self,
        kb_id: str,
        document_ids: list[str],
        *,
        job: JobRecord,
        delete_source_file: bool = False,
        delete_artifacts: bool = False,
    ) -> tuple[list[DocumentRecord], list[dict[str, Any]]]:
        self.assert_destructive_operation_supported("Batch document delete")
        async with self.kb_write_guard(kb_id) as record:
            claims = [
                (
                    document_id,
                    {
                        "pending_delete_job_id": job.id,
                        "delete_source_file": delete_source_file,
                        "delete_artifacts": delete_artifacts,
                    },
                )
                for document_id in document_ids
            ]
            return await self._metadata_store.claim_documents_deleting(
                record.id, claims
            )

    async def complete_delete(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        lightrag_result: dict[str, Any] | None = None,
        file_result: DocumentDeleteFileResult | None = None,
    ) -> DocumentRecord:
        self.assert_destructive_operation_supported("Document delete completion")
        async with self.kb_write_guard(kb_id) as record:
            document = await self._metadata_store.complete_document_delete(
                record.id,
                document_id,
                metadata_patch={
                    "pending_delete_job_id": None,
                    "current_delete_job_id": None,
                    "last_delete_job_id": job_id,
                    "last_deleted_at": utc_now_iso(),
                    "lightrag_delete_result": lightrag_result,
                    "file_delete_result": asdict(file_result) if file_result else None,
                },
            )
            await self._notify_agent_profile_dirty(
                record.id, document_id, "document_deleted"
            )
            return document

    async def _notify_agent_profile_dirty(
        self, kb_id: str, document_id: str, reason: str
    ) -> None:
        if self._agent_profile_dirty_callback is None:
            return
        try:
            await self._agent_profile_dirty_callback(kb_id, document_id, reason)
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "Agent profile dirty callback failed for KB '%s' doc '%s': %s",
                kb_id,
                document_id,
                exc,
            )

    async def fail_delete(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        error_code: str,
        error_message: str,
    ) -> DocumentRecord:
        async with self.kb_write_guard(kb_id) as record:
            return await self._metadata_store.fail_document_delete(
                record.id,
                document_id,
                error_code=error_code,
                error_message=error_message,
                metadata_patch={
                    "pending_delete_job_id": None,
                    "current_delete_job_id": None,
                    "last_failed_delete_job_id": job_id,
                },
            )

    def prepare_replacement_source(
        self, source: DocumentSourceInput
    ) -> DocumentReplacementSource:
        if not source.content:
            raise ValueError("Replacement document content cannot be empty")
        safe_name = _sanitize_source_name(source.source_name)
        return DocumentReplacementSource(
            source_name=safe_name,
            content=source.content,
            source_type=source.source_type,
            source_hash=_content_hash(source.content),
            content_type=source.content_type,
            size_bytes=len(source.content),
        )

    def _replacement_staging_path(
        self, workspace: str, document_id: str, job_id: str
    ) -> Path:
        """Deterministic on-disk location for replacement bytes staged at claim
        time so a durable worker can resume a ``replace`` job after a restart.

        Lives inside the document directory (same containment boundary as the
        source/artifacts) and is keyed by job id so concurrent replace attempts
        never collide. A leading dot keeps it out of artifact listings.
        """
        document_dir = (self._source_root / workspace / document_id).resolve(
            strict=False
        )
        return document_dir / f".replace-staging-{job_id}.bin"

    def _sync_staging_dir(self, workspace: str, batch_id: str) -> Path:
        """Deterministic directory for staged aggregate ``sync`` request bytes.

        Sync jobs are aggregate jobs (``document_id`` is ``None``), so their
        resumable state must live outside any one document directory until the
        worker decides whether each source creates, skips, or replaces a
        document. The batch id is generated before job creation and persisted in
        the job row, making this path reconstructable after restart.
        """
        workspace_dir = (self._source_root / workspace).resolve(strict=False)
        return workspace_dir / ".sync-staging" / batch_id

    def _sync_staging_path(
        self, workspace: str, batch_id: str, item_index: int, source_name: str
    ) -> Path:
        safe_name = _sanitize_source_name(source_name)
        return (
            self._sync_staging_dir(workspace, batch_id)
            / f"{item_index:04d}_{safe_name}"
        )

    async def stage_replacement_bytes(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        replacement: DocumentReplacementSource,
    ) -> str:
        """Persist replacement bytes to disk at claim time.

        The in-process replace task uses the in-memory ``replacement``; this
        staged copy exists purely so that if the process crashes mid-replace
        (orphan recovery → ``replace_failed`` → ``:retry`` → ``queued``), the
        durable worker can rebuild the ``DocumentReplacementSource`` from disk
        instead of needing the original request bytes.
        """
        self.assert_destructive_operation_supported("Document replace")
        async with self.kb_write_guard(kb_id) as record:
            document_dir = (self._source_root / record.workspace / document_id).resolve(
                strict=False
            )
            document_dir.mkdir(mode=0o700, parents=True, exist_ok=True)
            staging_path = self._replacement_staging_path(
                record.workspace, document_id, job_id
            )
            staging_path.write_bytes(replacement.content)
            return str(staging_path)

    async def load_staged_replacement(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        source_name: str,
        source_hash: str,
        content_type: str | None,
        size_bytes: int,
        source_type: SourceType,
    ) -> DocumentReplacementSource | None:
        """Rebuild a ``DocumentReplacementSource`` from staged bytes for worker
        resume. Returns ``None`` when the staging file is absent (e.g. the
        original request never staged, so the job is not worker-resumable).

        Raises ``ValueError`` when the staged bytes are present but their content
        hash no longer matches the ``source_hash`` persisted in the job payload
        (e.g. a truncated/corrupted staging file). The caller turns this into a
        clean ``replace_not_resumable`` failure instead of replaying the replace
        with wrong bytes — mirroring :meth:`load_staged_sync_source`.
        (``source_hash`` is always persisted for replace jobs created by the
        current code; the guard only skips verification for legacy payloads that
        never recorded one.)
        """
        self.assert_destructive_operation_supported("Document replace")
        record = await self._kb_service.get(kb_id)
        staging_path = self._replacement_staging_path(
            record.workspace, document_id, job_id
        )
        if not staging_path.is_file():
            return None
        content = staging_path.read_bytes()
        if source_hash:
            actual_hash = _content_hash(content)
            if actual_hash != source_hash:
                raise ValueError(
                    f"Staged replacement source hash mismatch for {source_name}: "
                    f"expected {source_hash}, got {actual_hash}"
                )
        return DocumentReplacementSource(
            source_name=source_name,
            content=content,
            source_type=source_type,
            source_hash=source_hash,
            content_type=content_type,
            size_bytes=size_bytes,
        )

    async def clear_staged_replacement(
        self, kb_id: str, document_id: str, *, job_id: str
    ) -> None:
        """Best-effort removal of the staged replacement bytes once the replace
        job reaches a terminal state."""
        try:
            record = await self._kb_service.get(kb_id)
        except Exception:  # noqa: BLE001 — never let cleanup break the caller
            return
        staging_path = self._replacement_staging_path(
            record.workspace, document_id, job_id
        )
        try:
            staging_path.unlink(missing_ok=True)
        except OSError:
            logger.warning(
                "Failed to remove staged replacement bytes at %s",
                staging_path,
            )

    async def stage_sync_source_bytes(
        self,
        kb_id: str,
        *,
        batch_id: str,
        item_index: int,
        source: DocumentSourceInput,
    ) -> str:
        """Persist one aggregate ``sync`` source before the job is queued.

        The in-process route can still use request-memory bytes, but the durable
        worker only depends on this staged copy plus the persisted job payload.
        """
        self.assert_destructive_operation_supported("Document sync")
        if not source.content:
            raise ValueError("Sync document content cannot be empty")
        async with self.kb_write_guard(kb_id) as record:
            staging_dir = self._sync_staging_dir(record.workspace, batch_id)
            staging_dir.mkdir(mode=0o700, parents=True, exist_ok=True)
            staging_path = self._sync_staging_path(
                record.workspace, batch_id, item_index, source.source_name
            )
            with staging_path.open("xb") as output:
                output.write(source.content)
                output.flush()
            return str(staging_path)

    async def load_staged_sync_source(
        self,
        kb_id: str,
        *,
        batch_id: str,
        item_index: int,
        source_name: str,
        content_type: str | None,
        metadata: dict[str, Any],
        expected_hash: str,
        source_type: SourceType,
    ) -> DocumentSourceInput | None:
        """Rebuild a sync source from staged bytes for worker resume."""
        self.assert_destructive_operation_supported("Document sync")
        record = await self._kb_service.get(kb_id)
        staging_path = self._sync_staging_path(
            record.workspace, batch_id, item_index, source_name
        )
        if not staging_path.is_file():
            return None
        content = staging_path.read_bytes()
        actual_hash = _content_hash(content)
        if actual_hash != expected_hash:
            raise ValueError(
                f"Staged sync source hash mismatch for {source_name}: "
                f"expected {expected_hash}, got {actual_hash}"
            )
        return DocumentSourceInput(
            source_name=source_name,
            content=content,
            source_type=source_type,
            content_type=content_type,
            metadata=metadata,
        )

    async def clear_staged_sync_sources(self, kb_id: str, *, batch_id: str) -> None:
        """Best-effort removal of staged aggregate sync bytes after terminal state."""
        try:
            record = await self._kb_service.get(kb_id)
        except Exception:  # noqa: BLE001 — never let cleanup break the caller
            return
        staging_dir = self._sync_staging_dir(record.workspace, batch_id)
        if staging_dir.exists():
            shutil.rmtree(staging_dir, ignore_errors=True)

    # ------------------------------------------------------------------
    # Object-backed replace/sync staging (Phase 3.2).
    #
    # These additive methods mirror the local-mode staging helpers above but
    # persist the staged bytes to immutable object-storage candidates instead
    # of local scratch files. They exist so that a durable worker can resume a
    # replace/sync job after the request process dies, without depending on a
    # local filesystem that may live under a different checkout/root on the
    # resuming worker.
    #
    # Replace staging reuses the *same* deterministic COW candidate key
    # (``_document_cow_source_object_key``) that the frozen Core Writer B1
    # state machine uploads to inside ``_upload_immutable_replacement_source``.
    # Because the key is deterministic for one source generation, the COW
    # commit's re-upload is idempotent: ``upload_file_if_absent`` finds the
    # object already present and returns ``created=False`` without overwriting
    # it. A crash between staging and COW commit therefore never loses the
    # bytes and never creates a second candidate.
    #
    # Sync staging uses a separate per-batch staging key: a sync item may
    # create a brand-new document whose ``document_id`` / source generation is
    # not yet known at request time, so the COW candidate key cannot be reused.
    # The worker downloads the staged sync bytes and re-drives the per-item
    # helper, which uploads the final source/artifact objects normally.
    # ------------------------------------------------------------------

    async def prepare_object_replace_staging(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        source_hash: str,
    ) -> tuple[str, str]:
        """Issue a fresh COW attempt token and its deterministic source generation id.

        Used by object-mode route staging so the request-time staging upload
        targets exactly the deterministic COW candidate key that the in-process
        execution and a durable worker resume will re-use. Returning the token
        here lets the route persist it in the job payload (``attempt_tokens``)
        before the in-process task runs, which makes the staging upload and the
        COW commit upload idempotent (same generation -> same key ->
        ``upload_file_if_absent`` returns ``created=False``).
        """

        if not self.object_authoritative:
            raise DocumentCowError(
                "Object replace staging identity requires object artifact mode"
            )
        record = await self._kb_service.get(kb_id)
        attempt_token = _new_document_attempt_token("replace")
        source_generation_id = document_source_generation_id(
            kb_id=kb_id,
            kb_generation=record.generation,
            document_id=document_id,
            job_id=job_id,
            attempt_token=attempt_token,
            source_hash=source_hash,
        )
        return attempt_token, source_generation_id

    async def stage_replacement_object(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        source_generation_id: str,
        replacement: DocumentReplacementSource,
    ) -> str:
        """Object-mode: stage replacement bytes to the deterministic COW candidate key.

        Writes the replacement bytes to an operation-scoped scratch file under
        the canonical input root (matching the frozen
        ``_upload_immutable_replacement_source`` pattern), uploads it via
        ``upload_file_if_absent(key=<COW candidate key>,
        expected_sha256=<source_hash>)``, removes the scratch file, and returns
        the staged object URI. The candidate key is the same key the COW commit
        will upload to, so the commit's upload is idempotent.

        Only valid in object artifact mode; local mode keeps using
        :meth:`stage_replacement_bytes`.
        """

        if not self.object_authoritative or self._object_storage is None:
            raise DocumentCowError(
                "Object replacement staging requires object artifact mode"
            )
        async with self.kb_write_guard(kb_id) as record:
            key = _document_cow_source_object_key(
                record.workspace,
                document_id,
                source_generation_id,
                replacement.source_name,
            )
            scratch_root = self._source_root / record.workspace
            scratch_root.mkdir(parents=True, exist_ok=True)
            scratch_path = scratch_root / f".replace-staging-{document_id}-{job_id}.tmp"
            try:
                scratch_path.write_bytes(replacement.content)
                uri, _created = await self._object_storage.upload_file_if_absent(
                    scratch_path,
                    key=key,
                    content_type=replacement.content_type,
                    expected_sha256=_extract_sha256_hex(replacement.source_hash),
                )
            finally:
                scratch_path.unlink(missing_ok=True)
            return uri

    async def load_staged_replacement_object(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        staging_object_uri: str,
        source_name: str,
        source_hash: str,
        content_type: str | None,
        size_bytes: int,
        source_type: SourceType,
    ) -> DocumentReplacementSource | None:
        """Rebuild a ``DocumentReplacementSource`` from a staged object URI.

        Downloads the staged object to an operation-scoped scratch file,
        verifies its SHA-256 matches ``source_hash``, and returns a
        ``DocumentReplacementSource`` carrying the materialized bytes, suitable
        for re-driving :meth:`execute_document_replace_cow`. Returns ``None``
        when the staged object is absent (e.g. the request never staged, so the
        job is not worker-resumable).

        The download is required because the frozen
        ``_upload_immutable_replacement_source`` writes the supplied
        ``replacement_content`` to a local scratch file and proves its checksum
        before the conditional PUT. Because the staged object already lives at
        the deterministic COW candidate key, that conditional PUT is a no-op
        (``created=False``) — the bytes are downloaded once for the proof and
        never re-uploaded. Raises ``ValueError`` when the downloaded bytes no
        longer match the persisted ``source_hash`` so the caller fails cleanly
        as ``replace_not_resumable`` instead of replaying wrong bytes.
        """

        if not self.object_authoritative or self._object_storage is None:
            raise DocumentCowError(
                "Object replacement staging requires object artifact mode"
            )
        record = await self._kb_service.get(kb_id)
        scratch_root = self._source_root / record.workspace
        scratch_root.mkdir(parents=True, exist_ok=True)
        scratch_path = (
            scratch_root / f".replace-staging-{document_id}-{job_id}.resume.tmp"
        )
        try:
            try:
                await self._object_storage.download_file(
                    staging_object_uri, scratch_path
                )
            except ObjectStorageNotFoundError:
                return None
            content = scratch_path.read_bytes()
        finally:
            scratch_path.unlink(missing_ok=True)
        if source_hash:
            actual_hex = _content_hash(content)
            expected_hex = _extract_sha256_hex(source_hash)
            if expected_hex is None:
                # Non-canonical persisted hash: compare verbatim (legacy guard).
                expected_hex = source_hash
            if actual_hex != expected_hex:
                raise ValueError(
                    f"Staged replacement object hash mismatch for {source_name}: "
                    f"expected {expected_hex}, got {actual_hex}"
                )
        return DocumentReplacementSource(
            source_name=source_name,
            content=content,
            source_type=source_type,
            source_hash=source_hash,
            content_type=content_type,
            size_bytes=size_bytes,
        )

    async def clear_staged_replacement_object(
        self,
        kb_id: str,
        document_id: str,  # noqa: ARG002 — kept for API symmetry with local mode
        *,
        job_id: str,  # noqa: ARG002 — kept for API symmetry with local mode
        staging_object_uri: str | None = None,  # noqa: ARG002 — see docstring
    ) -> None:
        """No-op for object-backed replace staging.

        The staged bytes live at the deterministic, immutable COW candidate key.
        On a successful replace that candidate becomes the document's current
        source pointer and is retained. On a failed/rolled-back replace the
        candidate is reclaimed through the existing cleanup-manifest
        infrastructure (the COW commit already enqueues the right manifests); a
        crash in the narrow window before the manifest is enqueued leaves a
        recoverable immutable orphan that Phase 3.2 orphan reconciliation
        (Writer O) handles. Object staging therefore never performs eager
        deletion here — doing so could delete the document's current source.
        """

        return None

    async def stage_sync_source_object(
        self,
        kb_id: str,
        *,
        batch_id: str,
        item_index: int,
        source: DocumentSourceInput,
    ) -> str:
        """Object-mode: stage one aggregate sync source to an immutable object.

        Sync items may create brand-new documents whose ``document_id`` and
        source generation are not known at request time, so the COW candidate
        key cannot be reused. Instead the bytes are staged under a deterministic
        per-batch, per-item staging key and the returned URI is persisted in the
        job payload. A durable worker downloads the bytes from that URI and
        re-drives the per-item sync helper, which uploads the final source/
        artifact objects normally.

        Only valid in object artifact mode; local mode keeps using
        :meth:`stage_sync_source_bytes`.
        """

        if not self.object_authoritative or self._object_storage is None:
            raise DocumentCowError("Object sync staging requires object artifact mode")
        if not source.content:
            raise ValueError("Sync document content cannot be empty")
        async with self.kb_write_guard(kb_id) as record:
            safe_name = _sanitize_source_name(source.source_name)
            key = (
                f"workspaces/{record.workspace}/sync-staging/{batch_id}/"
                f"{item_index:04d}/{safe_name}"
            )
            scratch_root = self._source_root / record.workspace
            scratch_root.mkdir(parents=True, exist_ok=True)
            scratch_path = (
                scratch_root / f".sync-staging-{batch_id}-{item_index:04d}.tmp"
            )
            try:
                scratch_path.write_bytes(source.content)
                uri, _created = await self._object_storage.upload_file_if_absent(
                    scratch_path,
                    key=key,
                    content_type=source.content_type,
                    expected_sha256=_content_hash(source.content),
                )
            finally:
                scratch_path.unlink(missing_ok=True)
            return uri

    async def load_staged_sync_source_object(
        self,
        kb_id: str,
        *,
        staging_object_uri: str,
        source_name: str,
        content_type: str | None,
        metadata: dict[str, Any],
        expected_hash: str,
        source_type: SourceType,
    ) -> DocumentSourceInput | None:
        """Rebuild one sync source from a staged object URI for worker resume.

        Downloads the staged object to an operation-scoped scratch file, verifies
        its SHA-256 matches ``expected_hash``, and returns a
        ``DocumentSourceInput`` carrying the materialized bytes. Returns ``None``
        when the staged object is absent. Raises ``ValueError`` on a checksum
        mismatch so the caller fails cleanly as ``sync_not_resumable`` instead of
        replaying wrong bytes.
        """

        if not self.object_authoritative or self._object_storage is None:
            raise DocumentCowError("Object sync staging requires object artifact mode")
        record = await self._kb_service.get(kb_id)
        scratch_root = self._source_root / record.workspace
        scratch_root.mkdir(parents=True, exist_ok=True)
        scratch_path = (
            scratch_root
            / f".sync-staging-resume-{_sanitize_source_name(source_name)}.tmp"
        )
        try:
            try:
                await self._object_storage.download_file(
                    staging_object_uri, scratch_path
                )
            except ObjectStorageNotFoundError:
                return None
            content = scratch_path.read_bytes()
        finally:
            scratch_path.unlink(missing_ok=True)
        actual_hex = _content_hash(content)
        expected_hex = _extract_sha256_hex(expected_hash)
        if expected_hex is None:
            expected_hex = expected_hash
        if actual_hex != expected_hex:
            raise ValueError(
                f"Staged sync source object hash mismatch for {source_name}: "
                f"expected {expected_hex}, got {actual_hex}"
            )
        return DocumentSourceInput(
            source_name=source_name,
            content=content,
            source_type=source_type,
            content_type=content_type,
            metadata=metadata,
        )

    async def clear_staged_sync_sources_object(
        self,
        kb_id: str,  # noqa: ARG002 — kept for API symmetry with local mode
        *,
        batch_id: str,  # noqa: ARG002 — see docstring
        staging_object_uris: Mapping[str, str] | None = None,  # noqa: ARG002
    ) -> None:
        """No-op for object-backed sync staging.

        Sync staging objects are immutable candidates under a per-batch staging
        prefix. They are reclaimed through the existing cleanup-manifest
        infrastructure (or Phase 3.2 orphan reconciliation for any crash-window
        orphans); eager deletion here could race with a concurrent/resuming
        worker that still needs the bytes. The local-mode cleanup is unaffected.
        """

        return None

    async def claim_replace(
        self,
        kb_id: str,
        document_id: str,
        *,
        job: JobRecord,
        replacement: DocumentReplacementSource,
        delete_source_file: bool = True,
        delete_artifacts: bool = True,
        delete_llm_cache: bool = False,
        auto_parse: bool = False,
        auto_index: bool = False,
        parser_engine: str | None = None,
        process_options: str | None = None,
        force_reparse: bool = False,
    ) -> DocumentRecord:
        self.assert_destructive_operation_supported("Document replace")
        async with self.kb_write_guard(kb_id) as record:
            return await self._metadata_store.claim_document_replacing(
                record.id,
                document_id,
                metadata_patch={
                    "pending_replace_job_id": job.id,
                    "replacement_source_name": replacement.source_name,
                    "replacement_source_hash": replacement.source_hash,
                    "delete_source_file": delete_source_file,
                    "delete_artifacts": delete_artifacts,
                    "delete_llm_cache": delete_llm_cache,
                    "auto_parse": auto_parse,
                    "auto_index": auto_index,
                    "parser_engine": parser_engine,
                    "process_options": process_options,
                    "force_reparse": force_reparse,
                },
            )

    async def replace_document_source(
        self,
        kb_id: str,
        document: DocumentRecord,
        *,
        job_id: str,
        replacement: DocumentReplacementSource,
        delete_source_file: bool = True,
        delete_artifacts: bool = True,
        lightrag_delete_result: dict[str, Any] | None = None,
    ) -> tuple[DocumentRecord, DocumentDeleteFileResult]:
        self.assert_destructive_operation_supported("Document replace")
        async with self.kb_write_guard(kb_id) as record:
            workspace_dir = (self._source_root / record.workspace).resolve(strict=False)
            document_dir = (workspace_dir / document.id).resolve(strict=False)
            try:
                document_dir.relative_to(workspace_dir)
            except ValueError as exc:
                raise ValueError(
                    "Document replacement path escapes workspace directory"
                ) from exc
            document_dir.mkdir(mode=0o700, parents=True, exist_ok=True)
            staging_path = document_dir / f".replace-{job_id}.tmp"
            staging_path.unlink(missing_ok=True)
            try:
                with staging_path.open("xb") as output:
                    output.write(replacement.content)
                    output.flush()

                file_result = await self.cleanup_document_files(
                    kb_id,
                    document,
                    delete_source_file=delete_source_file,
                    delete_artifacts=delete_artifacts,
                )
                if file_result.errors:
                    raise RuntimeError("; ".join(file_result.errors))

                target_path = _replacement_source_target(
                    document_dir, replacement.source_name, job_id
                )
                shutil.move(str(staging_path), str(target_path))
                source_object_uri = await self._persist_source_file(
                    record.workspace,
                    document.id,
                    target_path,
                    content_type=replacement.content_type,
                )
                replaced = await self._metadata_store.complete_document_replace(
                    record.id,
                    document.id,
                    source_name=replacement.source_name,
                    source_uri=str(target_path),
                    source_type=replacement.source_type,
                    source_hash=replacement.source_hash,
                    content_type=replacement.content_type,
                    size_bytes=replacement.size_bytes,
                    metadata_patch={
                        "pending_replace_job_id": None,
                        "current_replace_job_id": None,
                        "last_replace_job_id": job_id,
                        "last_replaced_at": utc_now_iso(),
                        "previous_lightrag_doc_id": document.lightrag_doc_id,
                        "lightrag_delete_result": lightrag_delete_result,
                        "file_replace_result": asdict(file_result),
                        **(
                            {"source_object_uri": source_object_uri}
                            if source_object_uri
                            else {}
                        ),
                    },
                )
                return replaced, file_result
            except Exception:
                staging_path.unlink(missing_ok=True)
                raise

    async def preflight_replace_cleanup(
        self,
        kb_id: str,
        document: DocumentRecord,
        *,
        delete_source_file: bool,
        delete_artifacts: bool,
    ) -> None:
        self.assert_destructive_operation_supported("Document replace")
        record = await self._kb_service.get(kb_id)
        workspace_dir = (self._source_root / record.workspace).resolve(strict=False)
        artifacts, _total = await self._metadata_store.list_document_artifacts(
            record.id, document.id, limit=200
        )
        _validate_document_cleanup_paths(
            workspace_dir,
            document,
            artifacts,
            delete_source_file=delete_source_file,
            delete_artifacts=delete_artifacts,
        )

    async def fail_replace(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        error_code: str,
        error_message: str,
        clear_index_metadata: bool = False,
        lightrag_delete_result: dict[str, Any] | None = None,
    ) -> DocumentRecord:
        async with self.kb_write_guard(kb_id) as record:
            return await self._metadata_store.fail_document_replace(
                record.id,
                document_id,
                error_code=error_code,
                error_message=error_message,
                clear_index_metadata=clear_index_metadata,
                metadata_patch={
                    "pending_replace_job_id": None,
                    "current_replace_job_id": None,
                    "last_failed_replace_job_id": job_id,
                    "lightrag_delete_result": lightrag_delete_result,
                },
            )

    async def cleanup_document_files(
        self,
        kb_id: str,
        document: DocumentRecord,
        *,
        delete_source_file: bool,
        delete_artifacts: bool,
    ) -> DocumentDeleteFileResult:
        self.assert_destructive_operation_supported("Document file cleanup")
        async with self.kb_write_guard(kb_id):
            return await self._cleanup_document_files_guarded(
                kb_id,
                document,
                delete_source_file=delete_source_file,
                delete_artifacts=delete_artifacts,
            )

    async def _cleanup_document_files_guarded(
        self,
        kb_id: str,
        document: DocumentRecord,
        *,
        delete_source_file: bool,
        delete_artifacts: bool,
    ) -> DocumentDeleteFileResult:
        record = await self._kb_service.get(kb_id)
        workspace_dir = (self._source_root / record.workspace).resolve(strict=False)
        # Canonical document directory: <source_root>/<workspace>/<document_id>.
        # Anchoring here (rather than trusting source_uri.parent) ensures both
        # source and artifact cleanup are contained to THIS document's dir, so a
        # crafted source_uri that lives inside the workspace but outside the doc
        # dir cannot escape the per-document boundary.
        document_dir = (workspace_dir / document.id).resolve(strict=False)
        result = DocumentDeleteFileResult()
        artifacts, _total = await self._metadata_store.list_document_artifacts(
            record.id, document.id, limit=200
        )
        source_path: Path | None = None
        if delete_source_file or delete_artifacts:
            try:
                source_path = _safe_document_path(
                    workspace_dir, document_dir, document.source_uri
                )
            except ValueError as exc:
                result.errors.append(f"source: {exc}")
                return result

        if delete_artifacts:
            for artifact in artifacts:
                try:
                    deleted_object = await self._delete_artifact_object(artifact)
                    if deleted_object:
                        result.deleted_objects.append(deleted_object)
                    artifact_path = _safe_document_path(
                        workspace_dir,
                        document_dir,
                        artifact.uri,
                    )
                    if not artifact_path.exists():
                        result.skipped.append(artifact.uri)
                        continue
                    _remove_path(artifact_path)
                    result.deleted_artifacts.append(str(artifact_path))
                except (OSError, ValueError, RuntimeError) as exc:
                    result.errors.append(f"artifact {artifact.id}: {exc}")

        if delete_source_file and source_path is not None:
            try:
                source_object_uri = document.metadata.get("source_object_uri")
                deleted_source_object = await self._delete_object_uri(source_object_uri)
                if deleted_source_object and isinstance(source_object_uri, str):
                    result.deleted_objects.append(source_object_uri)
                if source_path.exists():
                    _remove_path(source_path)
                else:
                    result.skipped.append(document.source_uri)
            except (OSError, ValueError, RuntimeError) as exc:
                result.errors.append(f"source: {exc}")
        return result

    async def create_parse_plan(
        self,
        kb_id: str,
        document_id: str,
        *,
        parser_engine: str | None = None,
        process_options: str | None = None,
        force_reparse: bool = False,
        auto_index: bool = False,
    ) -> DocumentParsePlan:
        async with self.kb_write_guard(kb_id):
            return await self._create_parse_plan_guarded(
                kb_id,
                document_id,
                parser_engine=parser_engine,
                process_options=process_options,
                force_reparse=force_reparse,
                auto_index=auto_index,
            )

    async def _create_parse_plan_guarded(
        self,
        kb_id: str,
        document_id: str,
        *,
        parser_engine: str | None = None,
        process_options: str | None = None,
        force_reparse: bool = False,
        auto_index: bool = False,
    ) -> DocumentParsePlan:
        record = await self._kb_service.get(kb_id)
        active_defaults = await self._active_parser_defaults_for_record(record)
        document = await self._metadata_store.get_document(record.id, document_id)
        source_name = _sanitize_source_name(document.source_name)

        engine, options = _resolve_parse_directives(
            Path(source_name),
            document,
            parser_engine=parser_engine,
            process_options=process_options,
            active_parser_engine=active_defaults.get("parser_engine"),
            active_process_options=active_defaults.get("process_options"),
        )
        lightrag_doc_id = document.lightrag_doc_id or compute_mdhash_id(
            str(document.source_uri), prefix="doc-"
        )
        parser_hash = _parser_hash(engine=engine, process_options=options)
        raw_object_refs = await self._raw_object_refs_for_plan(
            document, parser_engine=engine
        )
        raw_source_object_uri = document.metadata.get("source_object_uri")
        source_object_uri = (
            raw_source_object_uri
            if isinstance(raw_source_object_uri, str) and raw_source_object_uri
            else None
        )
        if self.object_authoritative and source_object_uri is not None:
            assert self._object_storage is not None
            self._object_storage.validate_document_file_uri(
                source_object_uri,
                workspace=document.workspace,
                document_id=document.id,
                namespace="source",
            )
        return DocumentParsePlan(
            document=document,
            source_name=source_name,
            source_object_uri=source_object_uri,
            raw_object_refs=raw_object_refs,
            parser_engine=engine,
            process_options=options,
            parser_hash=parser_hash,
            lightrag_doc_id=lightrag_doc_id,
            kb_generation=record.generation,
            expected_status=document.status,
            expected_source_hash=document.source_hash,
            expected_parser_hash=document.parser_hash,
            expected_current_parse_generation_id=document.metadata.get(
                "current_parse_generation_id"
            ),
            expected_current_sidecar_artifact_id=document.metadata.get(
                "current_sidecar_artifact_id"
            ),
            expected_current_blocks_artifact_id=document.metadata.get(
                "current_blocks_artifact_id"
            ),
            expected_index_hash=document.index_hash,
            force_reparse=force_reparse,
            auto_index=auto_index,
        )

    async def _raw_object_refs_for_plan(
        self, document: DocumentRecord, *, parser_engine: str
    ) -> tuple[DocumentRawObjectReference, ...]:
        if not self.object_authoritative or self._object_storage is None:
            return ()
        expected_names = set(_raw_directory_names(document.source_name, parser_engine))
        if not expected_names:
            return ()
        artifacts, _total = await self._metadata_store.list_document_artifacts(
            document.kb_id,
            document.id,
            artifact_type="raw_dir",
            limit=200,
            offset=0,
        )
        refs: list[DocumentRawObjectReference] = []
        seen_names: set[str] = set()
        for artifact in artifacts:
            if artifact.metadata.get("parse_engine") != parser_engine:
                continue
            object_prefix_uri = artifact.metadata.get("object_prefix_uri")
            if not isinstance(object_prefix_uri, str) or not object_prefix_uri:
                continue
            raw_name = artifact.metadata.get("raw_directory_name")
            if not isinstance(raw_name, str) or not raw_name:
                raw_name = Path(artifact.uri).name
            if raw_name not in expected_names or raw_name in seen_names:
                continue
            self._object_storage.validate_document_prefix_uri(
                object_prefix_uri,
                workspace=document.workspace,
                document_id=document.id,
                namespace="artifacts",
                artifact_id=artifact.id,
            )
            seen_names.add(raw_name)
            refs.append(
                DocumentRawObjectReference(
                    artifact_id=artifact.id,
                    object_prefix_uri=object_prefix_uri,
                    directory_name=raw_name,
                    checksum=artifact.checksum,
                )
            )
        return tuple(refs)

    async def create_batch_parse_plan(
        self,
        kb_id: str,
        document_ids: list[str],
        *,
        parser_engine: str | None = None,
        process_options: str | None = None,
        force_reparse: bool = False,
        auto_index: bool = False,
    ) -> DocumentBatchParsePlan:
        _validate_parse_request_directives(
            parser_engine=parser_engine, process_options=process_options
        )
        record = await self._kb_service.get(kb_id)
        plans: list[DocumentParsePlan] = []
        failures: list[dict[str, Any]] = []
        for document_id in document_ids:
            try:
                plan = await self.create_parse_plan(
                    record.id,
                    document_id,
                    parser_engine=parser_engine,
                    process_options=process_options,
                    force_reparse=force_reparse,
                    auto_index=auto_index,
                )
                plans.append(plan)
            except MetadataRecordNotFoundError as exc:
                failures.append(
                    _batch_parse_failure(
                        document_id,
                        error_code="document_not_found",
                        error_message=str(exc),
                    )
                )
            except ValueError as exc:
                failures.append(
                    _batch_parse_failure(
                        document_id,
                        error_code="invalid_parse_request",
                        error_message=str(exc),
                    )
                )
        return DocumentBatchParsePlan(
            batch_id=generate_track_id("batch"), plans=plans, failures=failures
        )

    async def mark_batch_parse_queued(
        self, kb_id: str, *, job: JobRecord, plans: list[DocumentParsePlan]
    ) -> list[DocumentRecord]:
        queued_documents, failures = await self.claim_batch_parse_queued(
            kb_id, job=job, plans=plans
        )
        if failures:
            failure = failures[0]
            if failure["error_code"] == "parse_job_active":
                raise ActiveDocumentParseJobError(
                    str(failure["document_id"]),
                    str(failure.get("existing_job_id") or "unknown"),
                )
            raise MetadataRecordNotFoundError(str(failure["error_message"]))
        return queued_documents

    async def claim_batch_parse_queued(
        self, kb_id: str, *, job: JobRecord, plans: list[DocumentParsePlan]
    ) -> tuple[list[DocumentRecord], list[dict[str, Any]]]:
        async with self.kb_write_guard(kb_id) as record:
            claims = [
                (
                    plan.document.id,
                    {
                        "pending_parse_job_id": job.id,
                        "pending_parse_batch_id": job.batch_id,
                        "pending_parser_hash": plan.parser_hash,
                        "pending_lightrag_doc_id": plan.lightrag_doc_id,
                        "parser_engine": plan.parser_engine,
                        "process_options": plan.process_options,
                        "force_reparse": plan.force_reparse,
                        "auto_index": plan.auto_index,
                    },
                    plan.expected_snapshot,
                    plan.claim_token,
                )
                for plan in plans
            ]
            (
                queued_documents,
                failures,
            ) = await self._metadata_store.claim_documents_parse_queued(
                record.id, claims
            )
            normalized_failures: list[dict[str, Any]] = []
            for failure in failures:
                normalized = failure
                if failure.get("error_code") == "document_snapshot_conflict":
                    document_id = failure.get("document_id")
                    if isinstance(document_id, str):
                        current = await self._metadata_store.get_document(
                            record.id, document_id
                        )
                        active_error = _active_document_job_error(current)
                        if active_error is not None:
                            normalized = {
                                **failure,
                                "error_code": _active_document_job_error_code(
                                    active_error
                                ),
                                "error_message": str(active_error),
                                "existing_job_id": active_error.existing_job_id,
                            }
                normalized_failures.append(normalized)
            plans_by_id = {plan.document.id: plan for plan in plans}
            for document in queued_documents:
                token = document.metadata.get("pending_parse_claim_token")
                if not isinstance(token, str) or not token:
                    raise DocumentLifecycleError(
                        "Parse claim did not persist an attempt token"
                    )
                plans_by_id[document.id].claim_token = token
                plans_by_id[document.id].job_id = job.id
            return queued_documents, normalized_failures

    async def mark_parse_queued(
        self, kb_id: str, document_id: str, *, job: JobRecord, plan: DocumentParsePlan
    ) -> DocumentRecord:
        async with self.kb_write_guard(kb_id) as record:
            try:
                document = await self._metadata_store.mark_document_parse_queued(
                    record.id,
                    document_id,
                    metadata_patch={
                        "pending_parse_job_id": job.id,
                        "pending_parser_hash": plan.parser_hash,
                        "pending_lightrag_doc_id": plan.lightrag_doc_id,
                        "parser_engine": plan.parser_engine,
                        "process_options": plan.process_options,
                        "force_reparse": plan.force_reparse,
                        "auto_index": plan.auto_index,
                    },
                    expected_snapshot=plan.expected_snapshot,
                    claim_token=plan.claim_token,
                )
            except DocumentSnapshotConflictError:
                current = await self._metadata_store.get_document(
                    record.id, document_id
                )
                active_error = _active_document_job_error(current)
                if active_error is not None:
                    raise active_error
                raise
            token = document.metadata.get("pending_parse_claim_token")
            if not isinstance(token, str) or not token:
                raise DocumentLifecycleError(
                    "Parse claim did not persist an attempt token"
                )
            plan.claim_token = token
            plan.job_id = job.id
            return document

    async def mark_parse_running(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        claim_token: str | None = None,
        plan: DocumentParsePlan | None = None,
    ) -> DocumentRecord:
        token = claim_token or (plan.claim_token if plan is not None else None)
        if self.object_authoritative and token is None:
            raise DocumentLifecycleError(
                "Object parse execution requires an explicit claim token"
            )
        async with self.kb_write_guard(kb_id) as record:
            document = await self._metadata_store.mark_document_parsing(
                record.id,
                document_id,
                metadata_patch={
                    "current_parse_job_id": job_id,
                    "parse_started_at": utc_now_iso(),
                },
                job_id=job_id if token is not None else None,
                claim_token=token,
            )
            resolved_token = document.metadata.get("current_parse_claim_token")
            if plan is not None and isinstance(resolved_token, str) and resolved_token:
                plan.claim_token = resolved_token
                plan.job_id = job_id
            return document

    async def materialize_parse_execution(
        self, plan: DocumentParsePlan
    ) -> DocumentParseExecution:
        canonical_document_root = self._canonical_document_root(plan.document)
        if not self.object_authoritative:
            source_path = await self._ensure_source_cached(plan.document)
            if not source_path.is_file():
                raise FileNotFoundError(
                    f"Document source not found: {plan.document.source_uri}"
                )
            return DocumentParseExecution(
                lease=None,
                scratch_document_root=source_path.parent,
                source_path=source_path,
                parsed_tree=source_path.parent / PARSED_DIR_NAME,
                canonical_document_root=canonical_document_root,
            )

        if not plan.claim_token:
            raise DocumentLifecycleError(
                "Object parse materialization requires a fenced attempt token"
            )

        if self._materializer is None:
            raise DocumentLifecycleError(
                "Object artifact mode has no configured materializer"
            )
        lease = self._materializer.create_lease()
        try:
            tree: MaterializedDocumentTree
            if plan.source_object_uri is not None:
                tree = await lease.materialize_document_source(
                    plan.source_object_uri,
                    workspace=plan.document.workspace,
                    document_id=plan.document.id,
                    source_name=plan.source_name,
                )
            else:
                fallback = self._object_mode_local_source_fallback(plan.document)
                tree = lease.link_document_source(
                    fallback, source_name=plan.source_name
                )

            if not plan.force_reparse:
                for raw_ref in plan.raw_object_refs:
                    restored_raw = await lease.materialize_document_prefix(
                        raw_ref.object_prefix_uri,
                        workspace=plan.document.workspace,
                        document_id=plan.document.id,
                        artifact_id=raw_ref.artifact_id,
                        tree=tree,
                        directory_name=raw_ref.directory_name,
                    )
                    if not _materialized_raw_cache_matches(
                        restored_raw, raw_ref.checksum
                    ):
                        logger.warning(
                            "Discarding raw parse cache artifact '%s' for document "
                            "'%s' because checksum verification failed",
                            raw_ref.artifact_id,
                            plan.document.id,
                        )
                        _remove_materialized_raw_cache(
                            restored_raw, document_root=tree.document_root
                        )
            return DocumentParseExecution(
                lease=lease,
                scratch_document_root=tree.document_root,
                source_path=tree.source_path,
                parsed_tree=tree.parsed_root,
                canonical_document_root=canonical_document_root,
            )
        except asyncio.CancelledError:
            lease.defer_cleanup()
            raise
        except Exception as exc:
            lease.cleanup()
            message = str(exc).replace(str(lease.path), "<artifact-materialization>")
            message = message.replace(".lightrag-scratch", "artifact-materialization")
            raise DocumentLifecycleError(message) from exc
        except BaseException:
            lease.cleanup()
            raise

    def _canonical_document_root(self, document: DocumentRecord) -> Path:
        input_root = self._source_root.expanduser().resolve(strict=False)
        workspace_root = (input_root / document.workspace).resolve(strict=False)
        document_root = (workspace_root / document.id).resolve(strict=False)
        if not workspace_root.is_relative_to(
            input_root
        ) or not document_root.is_relative_to(workspace_root):
            raise DocumentLifecycleError(
                "Canonical document root escapes configured INPUT_DIR"
            )
        return document_root

    def _object_mode_local_source_fallback(self, document: DocumentRecord) -> Path:
        input_root = self._source_root.expanduser().resolve(strict=False)
        candidate = Path(document.source_uri)
        if not candidate.is_absolute():
            candidate = input_root / candidate
        if candidate.is_symlink():
            raise DocumentLifecycleError(
                "Document source migration required: local fallback cannot be a symlink"
            )
        try:
            resolved = candidate.resolve(strict=True)
        except OSError as exc:
            raise DocumentLifecycleError(
                "Document source migration required: no source object reference and "
                "the canonical INPUT_DIR fallback is missing"
            ) from exc
        if not resolved.is_relative_to(input_root) or not resolved.is_file():
            raise DocumentLifecycleError(
                "Document source migration required: no source object reference and "
                "the local source is outside canonical INPUT_DIR"
            )
        scratch_root = (input_root / ".lightrag-scratch").resolve(strict=False)
        if resolved.is_relative_to(scratch_root):
            raise DocumentLifecycleError(
                "Document source migration required: scratch files cannot be used "
                "as a durable local fallback"
            )
        return resolved

    async def run_parse(
        self,
        rag: Any,
        plan: DocumentParsePlan,
        execution: DocumentParseExecution | None = None,
    ) -> dict[str, Any]:
        if execution is None:
            if self.object_authoritative:
                raise DocumentLifecycleError(
                    "Object parse execution must be materialized after a fenced claim"
                )
            execution = plan.execution
            if execution is None:
                execution = await self.materialize_parse_execution(plan)
                plan.execution = execution
        _verify_document_source_checksum(plan.document, execution.source_path)
        content_data = {
            "parse_format": FULL_DOCS_FORMAT_PENDING_PARSE,
            "parse_engine": plan.parser_engine,
            "process_options": plan.process_options,
            "force_reparse": plan.force_reparse,
            "archive_source_after_parse": False,
        }
        if self.object_authoritative and plan.job_id is not None:
            binding = _claimed_parse_artifact_binding(plan)
            content_data.update(
                {
                    "artifact_binding": binding.to_dict(),
                    "durable_file_path": canonicalize_pipeline_logical_filename(
                        plan.source_name
                    ),
                }
            )
        source_path = str(execution.source_path)
        if plan.parser_engine == PARSER_ENGINE_LEGACY:
            return await rag.parse_legacy(
                plan.lightrag_doc_id, source_path, content_data
            )
        if plan.parser_engine == PARSER_ENGINE_NATIVE:
            return await rag.parse_native(
                plan.lightrag_doc_id, source_path, content_data
            )
        if plan.parser_engine == PARSER_ENGINE_MINERU:
            return await rag.parse_mineru(
                plan.lightrag_doc_id, source_path, content_data
            )
        if plan.parser_engine == PARSER_ENGINE_DOCLING:
            return await rag.parse_docling(
                plan.lightrag_doc_id, source_path, content_data
            )
        raise ValueError(
            f"Unsupported parser engine for KB parse: {plan.parser_engine}"
        )

    async def finalize_parse_runtime_references(
        self,
        rag: Any,
        plan: DocumentParsePlan,
        execution: DocumentParseExecution,
        parsed_data: dict[str, Any] | None,
    ) -> None:
        """Re-emit the claimed parser row from the durable binding allowlist."""

        if not self.object_authoritative:
            return
        del execution, parsed_data
        full_docs = getattr(rag, "full_docs", None)
        get_by_id = getattr(full_docs, "get_by_id", None)
        if not callable(get_by_id):
            raise DocumentLifecycleError(
                "Parse binding finalizer requires durable full_docs storage"
            )
        get_record = cast(Callable[[str], Awaitable[Any]], get_by_id)
        existing = await get_record(plan.lightrag_doc_id)
        if not isinstance(existing, dict):
            raise DocumentLifecycleError(
                "Parse binding finalizer could not read the claimed full_docs row"
            )
        raw_binding = existing.get("artifact_binding")
        if not isinstance(raw_binding, dict):
            raise DocumentLifecycleError(
                "Parser first durable full_docs write is missing artifact_binding"
            )
        binding = PipelineArtifactBinding.from_mapping(
            raw_binding,
            expected_workspace=plan.document.workspace,
        )
        _validate_parse_artifact_binding(binding, plan, require_claimed=True)
        payload = _durable_binding_full_doc_payload(
            existing,
            binding=binding,
            file_path=plan.source_name,
        )
        assert_no_runtime_artifact_payload(
            {plan.lightrag_doc_id: payload},
            context="parse binding runtime finalizer",
        )
        await commit_pipeline_attempt_if_current(
            full_docs,
            plan.lightrag_doc_id,
            payload,
            expected_attempt_token=binding.claim_token,
            row_kind="full_docs",
        )

    async def commit_parse_artifact_binding(
        self,
        rag: Any,
        plan: DocumentParsePlan,
        result: DocumentParseResult,
    ) -> None:
        """Patch a claimed parse binding after metadata authority commits."""

        if not self.object_authoritative:
            return
        full_docs = getattr(rag, "full_docs", None)
        get_by_id = getattr(full_docs, "get_by_id", None)
        if not callable(get_by_id):
            raise DocumentLifecycleError(
                "Parse binding commit requires durable full_docs storage"
            )
        existing = await cast(Callable[[str], Awaitable[Any]], get_by_id)(
            plan.lightrag_doc_id
        )
        if not isinstance(existing, dict):
            raise DocumentLifecycleError(
                "Parse binding commit could not read the claimed full_docs row"
            )
        raw_binding = existing.get("artifact_binding")
        if not isinstance(raw_binding, dict):
            raise DocumentLifecycleError(
                "Parse binding commit could not read the claimed artifact_binding"
            )
        binding = PipelineArtifactBinding.from_mapping(
            raw_binding,
            expected_workspace=plan.document.workspace,
        )
        _validate_parse_artifact_binding(binding, plan, require_claimed=True)
        generation_id = result.document.metadata.get("current_parse_generation_id")
        if not isinstance(generation_id, str) or not generation_id:
            raise DocumentLifecycleError(
                "Parse binding commit could not resolve the committed generation"
            )
        committed = binding.committed(
            parse_generation_id=generation_id,
            index_hash=result.document.index_hash,
            sidecar_artifact_id=result.document.metadata.get(
                "current_sidecar_artifact_id"
            ),
            blocks_artifact_id=result.document.metadata.get(
                "current_blocks_artifact_id"
            ),
            raw_artifact_ids=sorted(
                artifact.id
                for artifact in result.artifacts
                if artifact.artifact_type == "raw_dir"
            ),
        )
        payload = _durable_binding_full_doc_payload(
            existing,
            binding=committed,
            file_path=plan.source_name,
        )
        assert_no_runtime_artifact_payload(
            {plan.lightrag_doc_id: payload},
            context="parse committed binding write",
        )
        await commit_pipeline_attempt_if_current(
            full_docs,
            plan.lightrag_doc_id,
            payload,
            expected_attempt_token=binding.claim_token,
            row_kind="full_docs",
        )

    async def complete_parse(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        plan: DocumentParsePlan,
        execution: DocumentParseExecution | None = None,
        parsed_data: dict[str, Any],
    ) -> DocumentParseResult:
        owns_compat_execution = execution is None
        if execution is None:
            if self.object_authoritative:
                raise DocumentLifecycleError(
                    "Object parse completion requires an explicit execution"
                )
            execution = plan.execution
            if execution is None:
                execution = await self.materialize_parse_execution(plan)
                plan.execution = execution

        try:
            current_document = await self._metadata_store.get_document(
                kb_id, document_id
            )
            resolved_token, phase = _resolve_service_attempt_owner(
                current_document,
                operation="parse",
                job_id=job_id,
                claim_token=plan.claim_token,
                strict=True,
            )
            if resolved_token is not None:
                plan.claim_token = resolved_token
            if phase != "current":
                await self.mark_parse_running(
                    kb_id,
                    document_id,
                    job_id=job_id,
                    claim_token=resolved_token,
                    plan=plan,
                )
            elif resolved_token is None:
                raise DocumentLifecycleError(
                    "Parse completion could not attach a token to a legacy running attempt"
                )
            generation_id = plan.claim_token
            if not isinstance(generation_id, str) or not generation_id:
                raise DocumentLifecycleError(
                    "Parse completion could not resolve a persisted attempt token"
                )
        except BaseException:
            if owns_compat_execution:
                execution.cleanup()
                if plan.execution is execution:
                    plan.execution = None
            raise

        try:
            async with self.kb_write_guard(kb_id) as record:
                pending_artifacts = _build_parse_artifacts(
                    plan,
                    execution,
                    parsed_data,
                    object_authoritative=self.object_authoritative,
                )
                pending_artifacts.extend(
                    _build_preview_artifacts(
                        plan,
                        execution,
                        parsed_data,
                        object_authoritative=self.object_authoritative,
                    )
                )
                for pending in pending_artifacts:
                    pending.record.metadata = {
                        **pending.record.metadata,
                        "parse_generation_id": generation_id,
                    }
                current_sidecar_artifact_id = next(
                    (
                        pending.record.id
                        for pending in pending_artifacts
                        if pending.record.artifact_type == "sidecar"
                    ),
                    None,
                )
                current_blocks_artifact_id = next(
                    (
                        pending.record.id
                        for pending in pending_artifacts
                        if pending.record.artifact_type == "blocks"
                    ),
                    None,
                )
                artifacts, uploaded_objects = await self._persist_parse_artifacts(
                    plan, pending_artifacts
                )
                blocks_path = _durable_blocks_path(
                    execution,
                    parsed_data.get("blocks_path"),
                    object_authoritative=self.object_authoritative,
                )
                try:
                    (
                        document,
                        created_artifacts,
                    ) = await self._metadata_store.complete_document_parse(
                        record.id,
                        document_id,
                        parser_hash=plan.parser_hash,
                        lightrag_doc_id=plan.lightrag_doc_id,
                        artifacts=artifacts,
                        retain_previous_artifacts=self.object_authoritative,
                        job_id=job_id,
                        claim_token=generation_id,
                        expected_snapshot=plan.expected_snapshot,
                        metadata_patch={
                            "last_parse_job_id": job_id,
                            "last_parsed_at": utc_now_iso(),
                            "parse_engine": plan.parser_engine,
                            "process_options": plan.process_options,
                            "parse_format": parsed_data.get(
                                "parse_format", FULL_DOCS_FORMAT_LIGHTRAG
                            ),
                            "blocks_path": blocks_path,
                            "artifact_count": len(artifacts),
                            "current_parse_generation_id": generation_id,
                            "current_build_generation_id": None,
                            "current_sidecar_artifact_id": current_sidecar_artifact_id,
                            "current_blocks_artifact_id": current_blocks_artifact_id,
                            "parse_stage_skipped": bool(
                                parsed_data.get("parse_stage_skipped")
                            ),
                        },
                    )
                except (Exception, asyncio.CancelledError) as commit_error:
                    return await self._reconcile_parse_artifact_commit_exception(
                        kb_id=record.id,
                        document_id=document_id,
                        artifacts=artifacts,
                        uploaded_objects=uploaded_objects,
                        job_id=job_id,
                        claim_token=generation_id,
                        parser_hash=plan.parser_hash,
                        lightrag_doc_id=plan.lightrag_doc_id,
                        current_sidecar_artifact_id=current_sidecar_artifact_id,
                        current_blocks_artifact_id=current_blocks_artifact_id,
                        commit_error=commit_error,
                    )
                return DocumentParseResult(
                    document=document, artifacts=created_artifacts
                )
        finally:
            if owns_compat_execution:
                execution.cleanup()
                if plan.execution is execution:
                    plan.execution = None

    async def _read_parse_artifact_commit_outcome(
        self,
        *,
        kb_id: str,
        document_id: str,
        artifacts: list[ArtifactRecord],
        job_id: str,
        claim_token: str,
        parser_hash: str,
        lightrag_doc_id: str,
        current_sidecar_artifact_id: str | None,
        current_blocks_artifact_id: str | None,
    ) -> MetadataCommitReconciliation[DocumentParseResult]:
        artifact_ids = [artifact.id for artifact in artifacts]
        (
            document,
            persisted_artifacts,
        ) = await self._metadata_store.get_document_and_artifacts_by_ids(
            kb_id,
            document_id,
            artifact_ids,
        )
        if document is None:
            return MetadataCommitReconciliation(
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="candidate_document_missing",
            )

        artifacts_complete = bool(
            len(persisted_artifacts) == len(artifacts)
            and set(persisted_artifacts) == set(artifact_ids)
            and all(
                _artifact_commit_candidate_matches(
                    artifact,
                    persisted_artifacts[artifact.id],
                )
                for artifact in artifacts
            )
        )
        metadata = document.metadata
        document_committed = bool(
            document.status == "parsed"
            and document.parser_hash == parser_hash
            and document.lightrag_doc_id == lightrag_doc_id
            and metadata.get("current_parse_generation_id") == claim_token
            and metadata.get("last_parse_job_id") == job_id
            and metadata.get("current_sidecar_artifact_id")
            == current_sidecar_artifact_id
            and metadata.get("current_blocks_artifact_id") == current_blocks_artifact_id
            and metadata.get("pending_parse_job_id") is None
            and metadata.get("pending_parse_claim_token") is None
            and metadata.get("current_parse_job_id") is None
            and metadata.get("current_parse_claim_token") is None
        )
        if artifacts_complete and document_committed:
            return MetadataCommitReconciliation(
                outcome=MetadataCommitOutcome.COMMITTED,
                value=DocumentParseResult(
                    document=document,
                    artifacts=[
                        persisted_artifacts[artifact_id] for artifact_id in artifact_ids
                    ],
                ),
                reason="candidate_document_and_artifacts_match",
            )

        candidate_pointer_visible = bool(
            metadata.get("current_parse_generation_id") == claim_token
            or (
                current_sidecar_artifact_id is not None
                and metadata.get("current_sidecar_artifact_id")
                == current_sidecar_artifact_id
            )
            or (
                current_blocks_artifact_id is not None
                and metadata.get("current_blocks_artifact_id")
                == current_blocks_artifact_id
            )
        )
        if not persisted_artifacts and not candidate_pointer_visible:
            return MetadataCommitReconciliation(
                outcome=MetadataCommitOutcome.ROLLED_BACK,
                reason="candidate_rows_and_pointers_absent",
            )
        return MetadataCommitReconciliation(
            outcome=MetadataCommitOutcome.UNKNOWN,
            reason=(
                "candidate_artifacts_partial_or_mismatched"
                if persisted_artifacts
                else "candidate_document_pointer_mismatch"
            ),
        )

    async def _reconcile_parse_artifact_commit_exception(
        self,
        *,
        kb_id: str,
        document_id: str,
        artifacts: list[ArtifactRecord],
        uploaded_objects: list[UploadedArtifactObject],
        job_id: str,
        claim_token: str,
        parser_hash: str,
        lightrag_doc_id: str,
        current_sidecar_artifact_id: str | None,
        current_blocks_artifact_id: str | None,
        commit_error: BaseException,
    ) -> DocumentParseResult:
        async def read_back() -> MetadataCommitReconciliation[DocumentParseResult]:
            return await self._read_parse_artifact_commit_outcome(
                kb_id=kb_id,
                document_id=document_id,
                artifacts=artifacts,
                job_id=job_id,
                claim_token=claim_token,
                parser_hash=parser_hash,
                lightrag_doc_id=lightrag_doc_id,
                current_sidecar_artifact_id=current_sidecar_artifact_id,
                current_blocks_artifact_id=current_blocks_artifact_id,
            )

        caller_cancelled = isinstance(commit_error, asyncio.CancelledError)
        readback_error: BaseException | None = None
        try:
            safe_result = await await_cancellation_safe_reconciliation(read_back)
            reconciliation = safe_result.value
            caller_cancelled = caller_cancelled or safe_result.caller_cancelled
        except asyncio.CancelledError as exc:
            readback_error = exc
            reconciliation = MetadataCommitReconciliation[DocumentParseResult](
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="readback_cancelled",
            )
        except Exception as exc:  # noqa: BLE001 - read failure means unknown
            readback_error = exc
            reconciliation = MetadataCommitReconciliation[DocumentParseResult](
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="readback_failed",
            )

        if (
            reconciliation.outcome is MetadataCommitOutcome.COMMITTED
            and reconciliation.value is not None
        ):
            if caller_cancelled:
                if isinstance(commit_error, asyncio.CancelledError):
                    raise commit_error
                raise asyncio.CancelledError() from commit_error
            return reconciliation.value

        if reconciliation.outcome is MetadataCommitOutcome.ROLLED_BACK:
            await self._compensate_uploaded_artifact_objects(uploaded_objects)
            if caller_cancelled:
                if isinstance(commit_error, asyncio.CancelledError):
                    raise commit_error
                raise asyncio.CancelledError() from commit_error
            raise commit_error

        artifact_ids = [artifact.id for artifact in artifacts]
        logger.warning(
            "metadata_commit_reconciliation outcome=unknown operation=%s "
            "document_id=%s candidate_artifact_ids=%s candidate_artifact_types=%s "
            "reason=%s commit_error_type=%s readback_error_type=%s",
            "parse_artifact_commit",
            document_id,
            artifact_ids,
            [artifact.artifact_type for artifact in artifacts],
            reconciliation.reason or "unknown",
            type(commit_error).__name__,
            type(readback_error).__name__ if readback_error is not None else None,
        )
        unknown_error = MetadataCommitOutcomeUnknownError(
            "parse_artifact_commit",
            candidate_document_ids=[document_id],
            candidate_job_id=job_id,
            candidate_artifact_ids=artifact_ids,
            candidate_artifact_types=[artifact.artifact_type for artifact in artifacts],
            reason=reconciliation.reason,
        )
        if caller_cancelled:
            if isinstance(commit_error, asyncio.CancelledError):
                raise commit_error from unknown_error
            raise asyncio.CancelledError() from unknown_error
        raise unknown_error from (readback_error or commit_error)

    async def fail_parse(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        plan: DocumentParsePlan,
        error_code: str,
        error_message: str,
    ) -> DocumentRecord:
        async with self.kb_write_guard(kb_id) as record:
            token = plan.claim_token
            phase: Literal["pending", "current"] | None = None
            if token is None:
                current_document = await self._metadata_store.get_document(
                    record.id, document_id
                )
                token, phase = _resolve_service_attempt_owner(
                    current_document,
                    operation="parse",
                    job_id=job_id,
                    claim_token=None,
                    strict=True,
                )
                if token is not None:
                    plan.claim_token = token
            if self.object_authoritative and token is None:
                raise DocumentLifecycleError(
                    "Object parse failure requires an explicit claim token"
                )
            metadata_patch = {
                "last_failed_parse_job_id": job_id,
                "last_failed_parser_hash": plan.parser_hash,
            }
            if phase == "pending":
                return await self._metadata_store.release_document_parse_if_owned(
                    record.id,
                    document_id,
                    job_id=job_id,
                    claim_token=token,
                    error_code=error_code,
                    error_message=error_message,
                    metadata_patch=metadata_patch,
                )
            return await self._metadata_store.fail_document_parse(
                record.id,
                document_id,
                error_code=error_code,
                error_message=error_message,
                job_id=job_id if token is not None or phase == "current" else None,
                claim_token=token,
                metadata_patch=metadata_patch,
            )

    async def release_parse_if_owned(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        plan: DocumentParsePlan,
        error_code: str,
        error_message: str,
    ) -> DocumentRecord:
        async with self.kb_write_guard(kb_id) as record:
            token = plan.claim_token
            if token is None:
                current_document = await self._metadata_store.get_document(
                    record.id, document_id
                )
                token, _phase = _resolve_service_attempt_owner(
                    current_document,
                    operation="parse",
                    job_id=job_id,
                    claim_token=None,
                    strict=False,
                )
                if token is not None:
                    plan.claim_token = token
            if self.object_authoritative and token is None:
                raise DocumentLifecycleError(
                    "Object parse release requires an explicit claim token"
                )
            return await self._metadata_store.release_document_parse_if_owned(
                record.id,
                document_id,
                job_id=job_id,
                claim_token=token,
                error_code=error_code,
                error_message=error_message,
                metadata_patch={
                    "last_failed_parse_job_id": job_id,
                    "last_failed_parser_hash": plan.parser_hash,
                },
            )

    async def list_document_artifacts(
        self,
        kb_id: str,
        document_id: str,
        *,
        artifact_type: str | None = None,
        limit: int = 50,
        offset: int = 0,
    ) -> tuple[list[ArtifactRecord], int]:
        record = await self._kb_service.get(kb_id)
        await self._metadata_store.get_document(record.id, document_id)
        return await self._metadata_store.list_document_artifacts(
            record.id,
            document_id,
            artifact_type=artifact_type,
            limit=limit,
            offset=offset,
        )

    async def get_document_preview_manifest(
        self, kb_id: str, document_id: str
    ) -> dict[str, Any]:
        record = await self._kb_service.get(kb_id)
        document = await self._metadata_store.get_document(record.id, document_id)
        preview_priority = {
            "preview_table_json": 0,
            "preview_text": 1,
            "preview_html": 2,
        }
        current_parse_generation = document.metadata.get("current_parse_generation_id")
        preview_artifacts: list[ArtifactRecord] = []
        for artifact_type in preview_priority:
            artifacts, _total = await self._metadata_store.list_document_artifacts(
                record.id,
                document_id,
                artifact_type=artifact_type,
                limit=1,
                offset=0,
            )
            if (
                artifacts
                and artifacts[0].metadata.get("preview") is True
                and (
                    not isinstance(current_parse_generation, str)
                    or artifacts[0].metadata.get("parse_generation_id")
                    == current_parse_generation
                )
            ):
                preview_artifacts.append(artifacts[0])

        variants: list[dict[str, Any]] = []
        for artifact in preview_artifacts:
            try:
                artifact_path, is_directory = self._manifest_artifact_path(
                    document, artifact
                )
            except (FileNotFoundError, ValueError, ObjectStorageError):
                continue
            media_type = _artifact_media_type(
                document, artifact, artifact_path, is_directory
            )
            variants.append(
                {
                    "kind": _preview_kind_for_artifact_type(artifact.artifact_type),
                    "artifact_id": artifact.id,
                    "artifact_type": artifact.artifact_type,
                    "media_type": media_type,
                    "size_bytes": _artifact_size_bytes(artifact, artifact_path),
                    "preview_url": _artifact_route_url(
                        record.id, document_id, artifact.id, suffix=":preview"
                    ),
                }
            )

        original_artifacts, _total = await self._metadata_store.list_document_artifacts(
            record.id,
            document_id,
            artifact_type="original",
            limit=1,
            offset=0,
        )
        original = original_artifacts[0] if original_artifacts else None
        if (
            original is not None
            and isinstance(current_parse_generation, str)
            and original.metadata.get("parse_generation_id") != current_parse_generation
        ):
            original = None
        fallback: dict[str, Any] | None = None
        if original is not None:
            try:
                original_path, original_is_dir = self._manifest_artifact_path(
                    document, original
                )
                original_media_type = _artifact_media_type(
                    document, original, original_path, original_is_dir
                )
                original_preview_kind = _inline_preview_kind_for_media_type(
                    original_media_type
                )
                if original_preview_kind is not None:
                    variants.append(
                        {
                            "kind": original_preview_kind,
                            "artifact_id": original.id,
                            "artifact_type": original.artifact_type,
                            "media_type": original_media_type,
                            "size_bytes": _artifact_size_bytes(original, original_path),
                            "preview_url": _artifact_route_url(
                                record.id,
                                document_id,
                                original.id,
                                suffix=":preview",
                            ),
                        }
                    )
                fallback = {
                    "artifact_id": original.id,
                    "artifact_type": original.artifact_type,
                    "media_type": original_media_type,
                    "size_bytes": _artifact_size_bytes(original, original_path),
                    "download_url": _artifact_route_url(
                        record.id, document_id, original.id, suffix=":download"
                    ),
                }
            except (FileNotFoundError, ValueError, ObjectStorageError):
                fallback = None

        return {
            "document_id": document.id,
            "source_name": document.source_name,
            "source_content_type": document.content_type,
            "status": document.status,
            "preferred": variants[0] if variants else None,
            "variants": variants,
            "fallback": fallback,
        }

    def _manifest_artifact_path(
        self, document: DocumentRecord, artifact: ArtifactRecord
    ) -> tuple[Path, bool]:
        if not self.object_authoritative:
            return _resolve_artifact_path(self._source_root, document, artifact)
        if self._object_storage is None:
            raise ValueError("Object storage is not enabled")
        object_prefix_uri = artifact.metadata.get("object_prefix_uri")
        if isinstance(object_prefix_uri, str) and object_prefix_uri:
            self._object_storage.validate_document_prefix_uri(
                object_prefix_uri,
                workspace=document.workspace,
                document_id=document.id,
                namespace="artifacts",
                artifact_id=artifact.id,
            )
            return Path(
                _artifact_logical_filename(document, artifact, allow_uri=False)
            ), True
        if artifact.metadata.get("is_directory"):
            raise FileNotFoundError(
                f"Directory artifact '{artifact.id}' requires migration or reparse"
            )
        object_uri = artifact.metadata.get("object_uri")
        if not isinstance(object_uri, str) or not object_uri:
            raise FileNotFoundError(
                f"Artifact '{artifact.id}' requires migration or reparse"
            )
        namespace = "source" if artifact.artifact_type == "original" else "artifacts"
        self._object_storage.validate_document_file_uri(
            object_uri,
            workspace=document.workspace,
            document_id=document.id,
            namespace=namespace,
            artifact_id=None if namespace == "source" else artifact.id,
        )
        return Path(
            _artifact_logical_filename(document, artifact, allow_uri=False)
        ), False

    async def get_document_artifact(
        self, kb_id: str, document_id: str, artifact_id: str
    ) -> ArtifactRecord:
        record = await self._kb_service.get(kb_id)
        await self._metadata_store.get_document(record.id, document_id)
        return await self._metadata_store.get_document_artifact(
            record.id, document_id, artifact_id
        )

    async def get_document_artifact_file(
        self, kb_id: str, document_id: str, artifact_id: str
    ) -> ArtifactFileResult:
        async with self.kb_write_guard(kb_id) as record:
            document = await self._metadata_store.get_document(record.id, document_id)
            artifact = await self._metadata_store.get_document_artifact(
                record.id, document_id, artifact_id
            )
            if self.object_authoritative:
                return await self._materialize_artifact_file(document, artifact)
            artifact_path, is_directory = _resolve_artifact_path(
                self._source_root, document, artifact
            )
            artifact_path = await self._ensure_artifact_cached(
                document, artifact, artifact_path
            )
            is_directory = artifact_path.is_dir()
            media_type = _artifact_media_type(
                document, artifact, artifact_path, is_directory
            )
            return ArtifactFileResult(
                artifact=artifact,
                path=artifact_path,
                filename=artifact_path.name + (".zip" if is_directory else ""),
                media_type=media_type,
                is_directory=is_directory,
            )

    async def _materialize_artifact_file(
        self, document: DocumentRecord, artifact: ArtifactRecord
    ) -> ArtifactFileResult:
        if self._materializer is None or self._object_storage is None:
            raise DocumentLifecycleError(
                "Object artifact download requires a materializer"
            )
        lease = self._materializer.create_lease()
        logical_filename = _artifact_logical_filename(
            document, artifact, allow_uri=False
        )
        object_prefix_uri = artifact.metadata.get("object_prefix_uri")
        object_uri = artifact.metadata.get("object_uri")
        try:
            if artifact.metadata.get("is_directory") and not (
                isinstance(object_prefix_uri, str) and object_prefix_uri
            ):
                raise FileNotFoundError(
                    f"Directory artifact '{artifact.id}' requires migration or reparse"
                )
            if isinstance(object_prefix_uri, str) and object_prefix_uri:
                materialized_path = await lease.materialize_prefix(
                    object_prefix_uri,
                    workspace=document.workspace,
                    document_id=document.id,
                    namespace="artifacts",
                    artifact_id=artifact.id,
                    target_name=_artifact_materialization_target_name(
                        artifact, logical_filename, is_directory=True
                    ),
                )
                is_directory = True
            elif isinstance(object_uri, str) and object_uri:
                namespace = (
                    "source" if artifact.artifact_type == "original" else "artifacts"
                )
                materialized_path = await lease.materialize_file(
                    object_uri,
                    workspace=document.workspace,
                    document_id=document.id,
                    namespace=namespace,
                    artifact_id=None if namespace == "source" else artifact.id,
                    target_name=_artifact_materialization_target_name(
                        artifact, logical_filename, is_directory=False
                    ),
                )
                is_directory = False
            else:
                raise FileNotFoundError(
                    f"Artifact '{artifact.id}' requires migration or reparse"
                )
            return ArtifactFileResult(
                artifact=artifact,
                path=materialized_path,
                filename=logical_filename + (".zip" if is_directory else ""),
                media_type=_artifact_media_type(
                    document,
                    artifact,
                    Path(logical_filename),
                    is_directory,
                ),
                is_directory=is_directory,
                lease=lease,
            )
        except asyncio.CancelledError:
            lease.defer_cleanup()
            raise
        except Exception as exc:
            message = (
                str(exc)
                .replace(str(lease.path), "<artifact-materialization>")
                .replace(".lightrag-scratch", "artifact-materialization")
            )
            lease.cleanup()
            raise DocumentLifecycleError(message) from exc
        except BaseException:
            lease.cleanup()
            raise

    async def get_document_artifact_download_url(
        self,
        kb_id: str,
        document_id: str,
        artifact_id: str,
        *,
        expires_in_seconds: int = 3600,
    ) -> ArtifactDownloadUrlResult:
        if expires_in_seconds <= 0:
            raise ValueError("expires_in_seconds must be positive")
        if self._object_storage is None:
            raise ValueError("Object storage is not enabled")
        record = await self._kb_service.get(kb_id)
        document = await self._metadata_store.get_document(record.id, document_id)
        artifact = await self._metadata_store.get_document_artifact(
            record.id, document_id, artifact_id
        )
        object_uri = artifact.metadata.get("object_uri")
        object_prefix_uri = artifact.metadata.get("object_prefix_uri")
        if self.object_authoritative:
            is_directory = bool(
                artifact.metadata.get("is_directory") or object_prefix_uri
            )
            artifact_path = Path(
                _artifact_logical_filename(document, artifact, allow_uri=False)
            )
        else:
            artifact_path, is_directory = _resolve_artifact_path(
                self._source_root, document, artifact
            )
        if is_directory or artifact.metadata.get("is_directory") or object_prefix_uri:
            raise ValueError(
                "Presigned download URLs are only available for file artifacts"
            )
        if not isinstance(object_uri, str) or not object_uri:
            raise ValueError("Artifact is not stored as an object-storage file")
        try:
            if self.object_authoritative:
                namespace = (
                    "source" if artifact.artifact_type == "original" else "artifacts"
                )
                self._object_storage.validate_document_file_uri(
                    object_uri,
                    workspace=document.workspace,
                    document_id=document.id,
                    namespace=namespace,
                    artifact_id=None if namespace == "source" else artifact.id,
                )
            else:
                self._object_storage.validate_document_file_uri(
                    object_uri,
                    workspace=document.workspace,
                    document_id=document.id,
                )
            url = await self._object_storage.presign_download_url(
                object_uri, expires_in_seconds=expires_in_seconds
            )
        except ObjectStorageError as exc:
            raise RuntimeError(str(exc)) from exc
        return ArtifactDownloadUrlResult(
            artifact=artifact,
            url=url,
            object_uri=object_uri,
            expires_in_seconds=expires_in_seconds,
            filename=_artifact_logical_filename(
                document,
                artifact,
                allow_uri=not self.object_authoritative,
            ),
            media_type=_artifact_media_type(document, artifact, artifact_path, False),
        )

    async def _documents_for_job(
        self, kb_id: str, job: JobRecord
    ) -> list[DocumentRecord]:
        document_ids = job.payload.get("document_ids")
        if isinstance(document_ids, list) and all(
            isinstance(document_id, str) for document_id in document_ids
        ):
            return await self._metadata_store.get_documents_by_ids(kb_id, document_ids)
        if job.batch_id:
            return await self._metadata_store.list_documents_by_batch_id(
                kb_id, job.batch_id
            )
        return []

    @staticmethod
    def _cleanup_saved_sources(saved_paths: list[Path], saved_dirs: list[Path]) -> None:
        for path in saved_paths:
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass
        for directory in saved_dirs:
            try:
                directory.rmdir()
            except OSError:
                pass

    async def _cleanup_saved_objects(self, object_uris: list[str]) -> None:
        for object_uri in object_uris:
            await self._delete_object_uri(object_uri)

    async def _persist_source_file(
        self,
        workspace: str,
        document_id: str,
        path: Path,
        *,
        content_type: str | None,
    ) -> str | None:
        if self._object_storage is None:
            return None
        key = f"workspaces/{workspace}/documents/{document_id}/source/{path.name}"
        object_uri = await self._object_storage.upload_file(
            path, key=key, content_type=content_type
        )
        try:
            self._object_storage.validate_document_file_uri(
                object_uri,
                workspace=workspace,
                document_id=document_id,
                namespace="source",
            )
        except BaseException:
            try:
                await self._delete_object_uri(object_uri)
            except Exception as cleanup_error:  # noqa: BLE001
                logger.warning(
                    "Failed to compensate invalid source object '%s': %s",
                    object_uri,
                    cleanup_error,
                )
            raise
        return object_uri

    async def _persist_parse_artifacts(
        self, plan: DocumentParsePlan, artifacts: list[PendingArtifact]
    ) -> tuple[list[ArtifactRecord], list[UploadedArtifactObject]]:
        if self._object_storage is None:
            return [pending.record for pending in artifacts], []

        uploaded: list[UploadedArtifactObject] = []
        records: list[ArtifactRecord] = []
        try:
            for pending in artifacts:
                artifact = pending.record
                path = pending.runtime_path
                metadata = dict(artifact.metadata)
                if artifact.artifact_type == "original" and plan.source_object_uri:
                    self._object_storage.validate_document_file_uri(
                        plan.source_object_uri,
                        workspace=plan.document.workspace,
                        document_id=plan.document.id,
                        namespace="source",
                    )
                    metadata["object_uri"] = plan.source_object_uri
                elif artifact.artifact_type == "original" and self.object_authoritative:
                    # A contained legacy local fallback is intentionally not
                    # promoted into a source object as a parse-time side effect.
                    # Phase 3 migration owns creation of that durable source ref.
                    pass
                elif pending.is_directory:
                    object_prefix_uri = await self._object_storage.upload_directory(
                        path,
                        prefix=_artifact_object_prefix(plan.document, artifact, path),
                    )
                    metadata["object_prefix_uri"] = object_prefix_uri
                    uploaded.append(
                        UploadedArtifactObject(uri=object_prefix_uri, is_prefix=True)
                    )
                    self._object_storage.validate_document_prefix_uri(
                        object_prefix_uri,
                        workspace=plan.document.workspace,
                        document_id=plan.document.id,
                        namespace="artifacts",
                        artifact_id=artifact.id,
                    )
                elif path.is_file():
                    object_uri = await self._object_storage.upload_file(
                        path,
                        key=_artifact_object_key(plan.document, artifact, path),
                        content_type=(
                            plan.document.content_type
                            if artifact.artifact_type == "original"
                            else None
                        ),
                    )
                    metadata["object_uri"] = object_uri
                    uploaded.append(
                        UploadedArtifactObject(uri=object_uri, is_prefix=False)
                    )
                    self._object_storage.validate_document_file_uri(
                        object_uri,
                        workspace=plan.document.workspace,
                        document_id=plan.document.id,
                        namespace="artifacts",
                        artifact_id=artifact.id,
                    )
                elif self.object_authoritative:
                    raise FileNotFoundError(
                        f"Parse artifact runtime file not found: {path}"
                    )
                artifact.metadata = metadata
                records.append(artifact)
        except BaseException:
            await self._compensate_uploaded_artifact_objects(uploaded)
            raise
        return records, uploaded

    async def _compensate_uploaded_artifact_objects(
        self, uploaded: list[UploadedArtifactObject]
    ) -> None:
        for uploaded_object in reversed(uploaded):
            try:
                if uploaded_object.is_prefix:
                    await self._delete_object_prefix(uploaded_object.uri)
                else:
                    await self._delete_object_uri(uploaded_object.uri)
            except Exception as exc:  # noqa: BLE001 - preserve the commit error
                logger.warning(
                    "Failed to compensate uncommitted parse object '%s': %s",
                    uploaded_object.uri,
                    exc,
                )

    async def _ensure_source_cached(self, document: DocumentRecord) -> Path:
        source_path = Path(document.source_uri)
        if source_path.is_file():
            return source_path
        object_uri = document.metadata.get("source_object_uri")
        if (
            self._object_storage is None
            or not isinstance(object_uri, str)
            or not object_uri
        ):
            return source_path
        await self._object_storage.download_file(object_uri, source_path)
        return source_path

    async def _ensure_artifact_cached(
        self, document: DocumentRecord, artifact: ArtifactRecord, artifact_path: Path
    ) -> Path:
        if artifact_path.exists():
            return artifact_path
        if self._object_storage is None:
            return artifact_path
        object_uri = artifact.metadata.get("object_uri")
        if isinstance(object_uri, str) and object_uri:
            await self._object_storage.download_file(object_uri, artifact_path)
            return artifact_path
        object_prefix_uri = artifact.metadata.get("object_prefix_uri")
        if isinstance(object_prefix_uri, str) and object_prefix_uri:
            artifact_path.mkdir(parents=True, exist_ok=True)
            await self._object_storage.download_prefix(object_prefix_uri, artifact_path)
            return artifact_path
        return artifact_path

    async def _delete_artifact_object(self, artifact: ArtifactRecord) -> str | None:
        object_prefix_uri = artifact.metadata.get("object_prefix_uri")
        if isinstance(object_prefix_uri, str) and object_prefix_uri:
            deleted = await self._delete_object_prefix(object_prefix_uri)
            return object_prefix_uri if deleted else None
        object_uri = artifact.metadata.get("object_uri")
        if isinstance(object_uri, str) and object_uri:
            deleted = await self._delete_object_uri(object_uri)
            return object_uri if deleted else None
        return None

    async def _delete_object_uri(self, object_uri: object) -> bool:
        if (
            self._object_storage is None
            or not isinstance(object_uri, str)
            or not object_uri
        ):
            return False
        try:
            return await self._object_storage.delete_uri(object_uri)
        except ObjectStorageError as exc:
            raise RuntimeError(str(exc)) from exc

    async def _delete_object_prefix(self, object_uri: object) -> bool:
        if (
            self._object_storage is None
            or not isinstance(object_uri, str)
            or not object_uri
        ):
            return False
        try:
            return bool(await self._object_storage.delete_prefix(object_uri))
        except ObjectStorageError as exc:
            raise RuntimeError(str(exc)) from exc

    # ------------------------------------------------------------------
    # Object-authoritative document copy-on-write execution (Phase 3.1).
    #
    # One shared implementation covers route and worker callers.  The Store A
    # COW metadata APIs own all destructive document/manifest state; this
    # service only orchestrates the immutable upload proof, exact manifest
    # preparation, the cancellation-safe commit reconciliation, and the
    # idempotent engine delete.  No direct object/local byte cleanup is ever
    # performed in the COW paths.
    # ------------------------------------------------------------------

    def _cow_grace_windows(
        self, *, now: datetime, replacement: bool
    ) -> tuple[datetime, datetime, datetime]:
        config = self._artifact_cleanup_config
        grace = (
            config.replacement_grace_seconds
            if replacement
            else config.staging_grace_seconds
        )
        delete_after = now + timedelta(seconds=grace)
        cleanup_deadline = delete_after + timedelta(seconds=config.cleanup_slo_seconds)
        audit_retain = now + timedelta(days=config.successful_audit_retention_days)
        return delete_after, cleanup_deadline, audit_retain

    async def _resolve_document_cow_target_authorities(
        self,
        claim: DocumentMutationClaimResult,
        *,
        retain_source: bool,
        retain_artifacts: bool,
    ) -> list[_DocumentCowTargetAuthority]:
        """Derive the exact Store A cleanup target set from one claim.

        Mirrors ``_document_mutation_cleanup_targets`` so the prepared
        manifests match Store A's authoritative target set exactly.
        """

        authorities: dict[str, _DocumentCowTargetAuthority] = {}
        source_uri = claim.old_source_object_uri
        source_generation = claim.old_source_generation_id
        if source_uri is not None:
            normalized = normalize_artifact_target_uri(source_uri)
            kind: Literal["object", "prefix"] = (
                "prefix" if normalized.endswith("/") else "object"
            )
            namespace: Literal["source", "legacy_source"] = (
                "source" if source_generation is not None else "legacy_source"
            )
            authority = _DocumentCowTargetAuthority(
                target_uri=normalized,
                target_kind=kind,
                target_namespace=namespace,
                artifact_id=None,
                source_generation_id=source_generation,
                disposition="retain" if retain_source else "delete",
            )
            authorities[authority.target_uri] = authority
        for artifact in claim.artifacts:
            metadata = artifact.metadata if isinstance(artifact.metadata, dict) else {}
            raw_targets: list[str] = []
            for key in ("object_uri", "object_prefix_uri"):
                value = metadata.get(key)
                if isinstance(value, str) and value:
                    raw_targets.append(value)
            if isinstance(artifact.uri, str) and "://" in artifact.uri:
                raw_targets.append(artifact.uri)
            for raw_target in raw_targets:
                normalized = normalize_artifact_target_uri(raw_target)
                if normalized in authorities:
                    continue
                artifact_kind: Literal["object", "prefix"] = (
                    "prefix" if normalized.endswith("/") else "object"
                )
                authority = _DocumentCowTargetAuthority(
                    target_uri=normalized,
                    target_kind=artifact_kind,
                    target_namespace="artifact",
                    artifact_id=artifact.id,
                    source_generation_id=None,
                    disposition="retain" if retain_artifacts else "delete",
                )
                authorities[normalized] = authority
        return sorted(authorities.values(), key=lambda item: item.target_uri)

    async def _inspect_document_cow_object_targets(
        self,
        authorities: list[_DocumentCowTargetAuthority],
    ) -> list[_DocumentCowManifestTarget]:
        """Capture comparable size/checksum/ETag/version for exact object targets.

        Prefix targets carry no invented checksum.  Already-absent objects are
        allowed (no expected evidence).  HEAD 403/unprovable blocks the whole
        group before any engine side effect.
        """

        if self._object_storage is None:
            raise DocumentCowError("Object COW requires enabled object storage")
        targets: list[_DocumentCowManifestTarget] = []
        for authority in authorities:
            if authority.target_kind != "object":
                targets.append(_DocumentCowManifestTarget(authority=authority))
                continue
            try:
                readback = await self._object_storage.inspect_object(
                    authority.target_uri
                )
            except ObjectStorageError as exc:
                raise DocumentCowManifestPreparationError(
                    "Object cleanup target could not be inspected"
                ) from exc
            if not readback.present or readback.stat is None:
                targets.append(_DocumentCowManifestTarget(authority=authority))
                continue
            stat = readback.stat
            targets.append(
                _DocumentCowManifestTarget(
                    authority=authority,
                    expected_size_bytes=stat.size,
                    expected_checksum=stat.checksum,
                    expected_etag=stat.etag,
                    expected_version_id=stat.version_id,
                )
            )
        return targets

    def _build_document_cow_manifests(
        self,
        operation: Literal["replace", "delete"],
        *,
        claim: DocumentMutationClaimResult,
        job_id: str,
        kb_generation: str,
        targets: list[_DocumentCowManifestTarget],
        now: datetime,
    ) -> tuple[str, tuple[ArtifactCleanupManifestRecord, ...]]:
        document = claim.document
        group_id = document_mutation_manifest_group_id(
            operation,
            kb_id=document.kb_id,
            kb_generation=kb_generation,
            document_id=document.id,
            job_id=job_id,
            attempt_token=claim.attempt_token,
            snapshot_digest=claim.snapshot_digest,
        )
        reason = "replace" if operation == "replace" else "document_delete"
        delete_after, cleanup_deadline, audit_retain = self._cow_grace_windows(
            now=now,
            replacement=(operation == "replace"),
        )
        seen_uris: set[str] = set()
        manifests: list[ArtifactCleanupManifestRecord] = []
        for target in targets:
            authority = target.authority
            if authority.target_uri in seen_uris:
                raise DocumentCowManifestPreparationError(
                    "Document COW produced duplicate cleanup targets"
                )
            seen_uris.add(authority.target_uri)
            disposition = authority.disposition
            status = "retained" if disposition == "retain" else "pending"
            idempotency_key = artifact_cleanup_idempotency_key(
                reason=reason,  # type: ignore[arg-type]
                kb_id=document.kb_id,
                kb_generation=kb_generation,
                workspace=document.workspace,
                document_id=document.id,
                artifact_id=authority.artifact_id,
                source_generation_id=authority.source_generation_id,
                target_kind=authority.target_kind,
                target_namespace=authority.target_namespace,
                target_uri=authority.target_uri,
            )
            manifest_id = _document_cow_manifest_id(
                operation,
                manifest_group_id=group_id,
                idempotency_key=idempotency_key,
            )
            manifests.append(
                ArtifactCleanupManifestRecord(
                    id=manifest_id,
                    idempotency_key=idempotency_key,
                    manifest_group_id=group_id,
                    kb_id=document.kb_id,
                    kb_generation=kb_generation,
                    workspace=document.workspace,
                    document_id=document.id,
                    artifact_id=authority.artifact_id,
                    source_generation_id=authority.source_generation_id,
                    origin_job_id=job_id,
                    origin_attempt_token=claim.attempt_token,
                    reason=reason,  # type: ignore[arg-type]
                    target_kind=authority.target_kind,
                    target_namespace=authority.target_namespace,
                    disposition=disposition,  # type: ignore[arg-type]
                    status=status,  # type: ignore[arg-type]
                    target_uri=authority.target_uri,
                    expected_size_bytes=target.expected_size_bytes,
                    expected_checksum=target.expected_checksum,
                    expected_etag=target.expected_etag,
                    expected_version_id=target.expected_version_id,
                    delete_after=delete_after,
                    cleanup_deadline_at=cleanup_deadline,
                    audit_retain_until=audit_retain,
                    next_attempt_at=delete_after,
                    created_at=now,
                    updated_at=now,
                )
            )
        return group_id, tuple(manifests)

    async def _prepare_document_cow_manifests(
        self,
        operation: Literal["replace", "delete"],
        claim: DocumentMutationClaimResult,
        *,
        job_id: str,
        kb_generation: str,
        retain_source: bool,
        retain_artifacts: bool,
    ) -> tuple[str, tuple[ArtifactCleanupManifestRecord, ...]]:
        authorities = await self._resolve_document_cow_target_authorities(
            claim,
            retain_source=retain_source,
            retain_artifacts=retain_artifacts,
        )
        targets = await self._inspect_document_cow_object_targets(authorities)
        return self._build_document_cow_manifests(
            operation,
            claim=claim,
            job_id=job_id,
            kb_generation=kb_generation,
            targets=targets,
            now=self._clock(),
        )

    async def _enqueue_orphan_reconcile_compensation(
        self,
        *,
        document: DocumentRecord,
        kb_generation: str,
        job_id: str,
        attempt_token: str,
        candidate_source_uri: str,
        source_generation_id: str,
    ) -> str:
        """Enqueue an orphan_reconcile/source manifest for a rolled-back candidate.

        Uses the accepted ``orphan_reconcile`` reason and ``source`` namespace.
        A crash between ``fail_document_replace_cow`` and this enqueue is a
        recoverable leak for Phase 3.2 orphan reconciliation; it is never a
        deletion or data-loss path.
        """

        normalized = normalize_artifact_target_uri(candidate_source_uri)
        now = self._clock()
        delete_after, cleanup_deadline, audit_retain = self._cow_grace_windows(
            now=now,
            replacement=True,
        )
        idempotency_key = artifact_cleanup_idempotency_key(
            reason="orphan_reconcile",
            kb_id=document.kb_id,
            kb_generation=kb_generation,
            workspace=document.workspace,
            document_id=document.id,
            artifact_id=None,
            source_generation_id=source_generation_id,
            target_kind="object",
            target_namespace="source",
            target_uri=normalized,
        )
        group_id = _document_cow_orphan_reconcile_group_id(
            kb_id=document.kb_id,
            kb_generation=kb_generation,
            document_id=document.id,
            job_id=job_id,
            attempt_token=attempt_token,
            candidate_uri=normalized,
        )
        manifest_id = _document_cow_manifest_id(
            "replace",
            manifest_group_id=group_id,
            idempotency_key=idempotency_key,
        )
        manifest = ArtifactCleanupManifestRecord(
            id=manifest_id,
            idempotency_key=idempotency_key,
            manifest_group_id=group_id,
            kb_id=document.kb_id,
            kb_generation=kb_generation,
            workspace=document.workspace,
            document_id=document.id,
            artifact_id=None,
            source_generation_id=source_generation_id,
            origin_job_id=job_id,
            origin_attempt_token=attempt_token,
            reason="orphan_reconcile",
            target_kind="object",
            target_namespace="source",
            disposition="delete",
            status="pending",
            target_uri=normalized,
            delete_after=delete_after,
            cleanup_deadline_at=cleanup_deadline,
            audit_retain_until=audit_retain,
            next_attempt_at=delete_after,
            created_at=now,
            updated_at=now,
        )
        await self._metadata_store.enqueue_artifact_cleanup_manifest(manifest)
        return manifest_id

    async def execute_document_replace_cow(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        kb_generation: str,
        new_source_type: str,
        new_source_name: str,
        new_source_uri: str,
        new_source_hash: str,
        new_content_type: str | None,
        new_size_bytes: int,
        replacement_content: bytes | Path,
        engine_delete: Callable[
            [str, str, str | None, str],
            Awaitable[dict[str, Any] | None],
        ],
        claim_token: str | None = None,
        expected_snapshot: str | None = None,
        retain_source: bool = False,
        retain_artifacts: bool = False,
        metadata_patch: dict[str, Any] | None = None,
    ) -> DocumentReplaceCowResult:
        """Execute one object-authoritative document replacement.

        Ordering guarantee: the new source pointer and exact cleanup manifest
        group commit atomically (Store A) BEFORE the previous LightRAG document
        is deleted.  A crash after commit leaves a re-driveable
        ``engine_cleanup_pending`` state; a crash between commit and engine
        delete never loses data.
        """

        if not self.object_authoritative:
            raise DocumentCowError("Document replace COW requires object artifact mode")
        if self._object_storage is None:
            raise DocumentCowError("Document replace COW requires object storage")
        async with self.kb_write_guard(kb_id, expected_generation=kb_generation):
            record = await assert_active_kb_generation(
                self._kb_service,
                self._metadata_store,
                kb_id,
                kb_generation,
            )
            (
                attempt_token,
                source_generation_id,
                claim,
                committed_doc,
            ) = await self._claim_or_resume_document_replace_cow(
                kb_id=kb_id,
                document_id=document_id,
                kb_generation=kb_generation,
                job_id=job_id,
                claim_token=claim_token,
                new_source_hash=new_source_hash,
            )
            new_source_object_uri = self._object_storage.object_uri_for_key(
                _document_cow_source_object_key(
                    record.workspace,
                    document_id,
                    source_generation_id,
                    new_source_name,
                )
            )
            if committed_doc is None:
                assert claim is not None  # exactly one of claim/committed_doc set
                commit_result = await self._commit_document_replace_cow_attempt(
                    claim=claim,
                    kb_id=kb_id,
                    document_id=document_id,
                    kb_generation=kb_generation,
                    job_id=job_id,
                    attempt_token=attempt_token,
                    new_source_type=new_source_type,
                    new_source_name=new_source_name,
                    new_source_uri=new_source_uri,
                    new_source_hash=new_source_hash,
                    new_content_type=new_content_type,
                    new_size_bytes=new_size_bytes,
                    new_source_object_uri=new_source_object_uri,
                    new_source_generation_id=source_generation_id,
                    replacement_content=replacement_content,
                    retain_source=retain_source,
                    retain_artifacts=retain_artifacts,
                    metadata_patch=metadata_patch,
                )
                committed_doc = commit_result.document
                manifest_group_id = commit_result.manifest_group_id
                manifest_ids = commit_result.manifest_ids
                pending = commit_result.pending_cleanup_count
                retained = commit_result.retained_cleanup_count
                blocked = commit_result.blocked_cleanup_count
            else:
                source_generation_id = str(
                    committed_doc.metadata.get("source_generation_id")
                    or source_generation_id
                )
                new_source_object_uri = str(
                    committed_doc.metadata.get("source_object_uri")
                    or new_source_object_uri
                )
                manifest_group_id = str(
                    committed_doc.metadata.get("last_replace_manifest_group_id") or ""
                )
                manifest_ids = tuple(
                    committed_doc.metadata.get("last_replace_manifest_ids") or []
                )
                pending, retained, blocked = await self._document_cow_cleanup_counts(
                    kb_id=kb_id,
                    manifest_group_id=manifest_group_id,
                )
            previous_lightrag_doc_id = committed_doc.metadata.get(
                "previous_lightrag_doc_id"
            )
            engine_identity = _document_cow_engine_identity(
                kb_id=kb_id,
                document_id=document_id,
                source_generation_id=source_generation_id,
            )
            try:
                engine_result = await self._run_document_cow_engine_delete(
                    engine_delete,
                    kb_id=kb_id,
                    document_id=document_id,
                    previous_lightrag_doc_id=previous_lightrag_doc_id,
                    engine_identity=engine_identity,
                )
            except Exception as exc:
                await self._metadata_store.record_document_replace_engine_cleanup_failure_cow(
                    kb_id,
                    document_id,
                    kb_generation=kb_generation,
                    job_id=job_id,
                    attempt_token=attempt_token,
                    source_object_uri=new_source_object_uri,
                    source_generation_id=source_generation_id,
                    manifest_group_id=manifest_group_id,
                    error_code="engine_cleanup_failed",
                    error_message=str(exc) or type(exc).__name__,
                )
                failure = await self._metadata_store.get_document(kb_id, document_id)
                result = DocumentReplaceCowResult(
                    document=failure,
                    attempt_token=attempt_token,
                    manifest_group_id=manifest_group_id,
                    manifest_ids=manifest_ids,
                    cleanup_pending_count=pending,
                    cleanup_retained_count=retained,
                    cleanup_blocked_count=blocked,
                    source_generation_id=source_generation_id,
                    phase="engine_cleanup_pending",
                    outcome="cleanup_pending",
                )
                error = DocumentCowEngineDeleteError(
                    "replace",
                    document_id=document_id,
                    job_id=job_id,
                    attempt_token=attempt_token,
                )
                error.result = result
                raise error from exc
            finalized = await self._metadata_store.finalize_document_replace_cow(
                kb_id,
                document_id,
                kb_generation=kb_generation,
                job_id=job_id,
                attempt_token=attempt_token,
                source_object_uri=new_source_object_uri,
                source_generation_id=source_generation_id,
                manifest_group_id=manifest_group_id,
            )
            if engine_result is not None or metadata_patch:
                finalized = await self._metadata_store.update_document(
                    kb_id,
                    document_id,
                    metadata_patch={
                        **(
                            {"lightrag_delete_result": engine_result}
                            if engine_result is not None
                            else {}
                        ),
                        **(metadata_patch or {}),
                    },
                )
            return DocumentReplaceCowResult(
                document=finalized,
                attempt_token=attempt_token,
                manifest_group_id=manifest_group_id,
                manifest_ids=manifest_ids,
                cleanup_pending_count=pending,
                cleanup_retained_count=retained,
                cleanup_blocked_count=blocked,
                source_generation_id=source_generation_id,
                phase="completed",
                outcome="completed",
            )

    async def _claim_or_resume_document_replace_cow(
        self,
        *,
        kb_id: str,
        document_id: str,
        kb_generation: str,
        job_id: str,
        claim_token: str | None,
        new_source_hash: str,
    ) -> tuple[str, str, DocumentMutationClaimResult | None, DocumentRecord | None]:
        """Claim a replace, or resume an already-committed replace attempt.

        Returns ``(attempt_token, source_generation_id, claim, committed_doc)``.
        Exactly one of ``claim`` / ``committed_doc`` is non-None: ``claim`` for
        the normal pre-commit path, ``committed_doc`` when a prior attempt
        already committed and the caller must skip straight to engine cleanup.
        """

        committed = await self._maybe_resume_committed_replace_cow(
            kb_id=kb_id,
            document_id=document_id,
            kb_generation=kb_generation,
            job_id=job_id,
            claim_token=claim_token,
        )
        if committed is not None:
            attempt_token = str(
                committed.metadata.get("current_replace_claim_token") or ""
            )
            generation = document_source_generation_id(
                kb_id=kb_id,
                kb_generation=kb_generation,
                document_id=document_id,
                job_id=job_id,
                attempt_token=attempt_token,
                source_hash=new_source_hash,
            )
            return attempt_token, generation, None, committed
        claim = await self._metadata_store.claim_document_replacing_cow(
            kb_id,
            document_id,
            kb_generation=kb_generation,
            job_id=job_id,
            claim_token=claim_token,
        )
        return (
            claim.attempt_token,
            document_source_generation_id(
                kb_id=kb_id,
                kb_generation=kb_generation,
                document_id=document_id,
                job_id=job_id,
                attempt_token=claim.attempt_token,
                source_hash=new_source_hash,
            ),
            claim,
            None,
        )

    async def _maybe_resume_committed_replace_cow(
        self,
        *,
        kb_id: str,
        document_id: str,
        kb_generation: str,
        job_id: str,
        claim_token: str | None,
    ) -> DocumentRecord | None:
        document = await self._metadata_store.get_document(kb_id, document_id)
        if document.metadata.get("replace_phase") != "engine_cleanup_pending":
            return None
        if document.metadata.get("current_replace_job_id") != job_id:
            return None
        attempt = document.metadata.get("current_replace_claim_token")
        if claim_token is not None and attempt != claim_token:
            return None
        if not isinstance(attempt, str) or not attempt:
            return None
        return document

    async def _commit_document_replace_cow_attempt(
        self,
        *,
        claim: DocumentMutationClaimResult,
        kb_id: str,
        document_id: str,
        kb_generation: str,
        job_id: str,
        attempt_token: str,
        new_source_type: str,
        new_source_name: str,
        new_source_uri: str,
        new_source_hash: str,
        new_content_type: str | None,
        new_size_bytes: int,
        new_source_object_uri: str,
        new_source_generation_id: str,
        replacement_content: bytes | Path,
        retain_source: bool,
        retain_artifacts: bool,
        metadata_patch: dict[str, Any] | None,
    ) -> DocumentMutationCommitResult:
        document = claim.document
        try:
            await self._upload_immutable_replacement_source(
                replacement_content=replacement_content,
                workspace=document.workspace,
                document_id=document_id,
                source_generation_id=new_source_generation_id,
                source_name=new_source_name,
                content_type=new_content_type,
                expected_sha256=_extract_sha256_hex(new_source_hash),
            )
            _group_id, manifests = await self._prepare_document_cow_manifests(
                "replace",
                claim,
                job_id=job_id,
                kb_generation=kb_generation,
                retain_source=retain_source,
                retain_artifacts=retain_artifacts,
            )
        except Exception:
            # A pre-commit side-effect failure (upload proof mismatch or
            # unprovable manifest target) must release the claim so the
            # document returns to its prior state. The candidate object, if
            # any, is an orphan for Phase 3.2 reconciliation; no commit ever
            # landed and no engine delete ran.
            await self._metadata_store.fail_document_replace_cow(
                kb_id,
                document_id,
                kb_generation=kb_generation,
                job_id=job_id,
                attempt_token=claim.attempt_token,
                error_code="replace_pre_commit_failed",
                error_message="Replacement pre-commit proof failed safely",
            )
            raise
        try:
            committed = await self._metadata_store.commit_document_replace_cow(
                kb_id,
                document_id,
                kb_generation=kb_generation,
                job_id=job_id,
                attempt_token=claim.attempt_token,
                expected_snapshot=claim.snapshot_digest,
                new_source_type=new_source_type,
                new_source_name=new_source_name,
                new_source_uri=new_source_uri or new_source_object_uri,
                new_source_hash=new_source_hash,
                new_content_type=new_content_type,
                new_size_bytes=new_size_bytes,
                new_source_object_uri=new_source_object_uri,
                new_source_generation_id=new_source_generation_id,
                metadata_patch=metadata_patch,
                manifests=manifests,
            )
        except (Exception, asyncio.CancelledError) as commit_error:
            return await self._reconcile_document_replace_cow_commit_exception(
                kb_id=kb_id,
                document_id=document_id,
                kb_generation=kb_generation,
                job_id=job_id,
                attempt_token=claim.attempt_token,
                expected_snapshot=claim.snapshot_digest,
                new_source_type=new_source_type,
                new_source_name=new_source_name,
                new_source_uri=new_source_uri,
                new_source_hash=new_source_hash,
                new_content_type=new_content_type,
                new_size_bytes=new_size_bytes,
                new_source_object_uri=new_source_object_uri,
                new_source_generation_id=new_source_generation_id,
                manifests=manifests,
                commit_error=commit_error,
            )
        return committed

    async def _upload_immutable_replacement_source(
        self,
        *,
        replacement_content: bytes | Path,
        workspace: str,
        document_id: str,
        source_generation_id: str,
        source_name: str,
        content_type: str | None,
        expected_sha256: str | None,
    ) -> tuple[str, bool]:
        assert self._object_storage is not None
        key = _document_cow_source_object_key(
            workspace,
            document_id,
            source_generation_id,
            source_name,
        )
        if isinstance(replacement_content, Path):
            local_path = replacement_content
        else:
            scratch_root = self._source_root / workspace
            scratch_root.mkdir(parents=True, exist_ok=True)
            local_path = (
                scratch_root / f".cow-replace-{document_id}-{source_generation_id}.tmp"
            )
            local_path.write_bytes(replacement_content)
        return await self._object_storage.upload_file_if_absent(
            local_path,
            key=key,
            content_type=content_type,
            expected_sha256=expected_sha256,
        )

    async def _reconcile_document_replace_cow_commit_exception(
        self,
        *,
        kb_id: str,
        document_id: str,
        kb_generation: str,
        job_id: str,
        attempt_token: str,
        expected_snapshot: str,
        new_source_type: str,
        new_source_name: str,
        new_source_uri: str,
        new_source_hash: str,
        new_content_type: str | None,
        new_size_bytes: int,
        new_source_object_uri: str,
        new_source_generation_id: str,
        manifests: tuple[ArtifactCleanupManifestRecord, ...],
        commit_error: BaseException,
    ) -> DocumentMutationCommitResult:
        caller_cancelled = isinstance(commit_error, asyncio.CancelledError)
        readback_error: BaseException | None = None
        try:
            safe = await await_cancellation_safe_reconciliation(
                lambda: self._metadata_store.reconcile_document_replace_cow_commit(
                    kb_id,
                    document_id,
                    kb_generation=kb_generation,
                    job_id=job_id,
                    attempt_token=attempt_token,
                    expected_snapshot=expected_snapshot,
                    new_source_type=new_source_type,
                    new_source_name=new_source_name,
                    new_source_uri=new_source_uri,
                    new_source_hash=new_source_hash,
                    new_content_type=new_content_type,
                    new_size_bytes=new_size_bytes,
                    new_source_object_uri=new_source_object_uri,
                    new_source_generation_id=new_source_generation_id,
                    manifests=manifests,
                )
            )
            reconciliation = safe.value
            caller_cancelled = caller_cancelled or safe.caller_cancelled
        except asyncio.CancelledError as exc:
            readback_error = exc
            reconciliation = MetadataCommitReconciliation[DocumentMutationCommitResult](
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="readback_cancelled",
            )
        except Exception as exc:  # noqa: BLE001
            readback_error = exc
            reconciliation = MetadataCommitReconciliation[DocumentMutationCommitResult](
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="readback_failed",
            )

        if (
            reconciliation.outcome is MetadataCommitOutcome.COMMITTED
            and reconciliation.value is not None
        ):
            return reconciliation.value
        if reconciliation.outcome is MetadataCommitOutcome.ROLLED_BACK:
            await self._metadata_store.fail_document_replace_cow(
                kb_id,
                document_id,
                kb_generation=kb_generation,
                job_id=job_id,
                attempt_token=attempt_token,
                error_code="replace_commit_rolled_back",
                error_message="Replacement metadata commit rolled back safely",
            )
            try:
                await self._enqueue_orphan_reconcile_compensation(
                    document=await self._metadata_store.get_document(
                        kb_id, document_id
                    ),
                    kb_generation=kb_generation,
                    job_id=job_id,
                    attempt_token=attempt_token,
                    candidate_source_uri=new_source_object_uri,
                    source_generation_id=new_source_generation_id,
                )
            except Exception as compensation_error:  # noqa: BLE001
                # A crash here is a recoverable leak for Phase 3.2 orphan
                # reconciliation; the rolled-back metadata is already durable.
                logger.warning(
                    "document_replace_cow compensation enqueue failed kb_id=%s "
                    "document_id=%s attempt_token=%s error=%s",
                    kb_id,
                    document_id,
                    attempt_token,
                    compensation_error,
                )
            if caller_cancelled:
                if isinstance(commit_error, asyncio.CancelledError):
                    raise commit_error
                raise asyncio.CancelledError() from commit_error
            raise DocumentCowRetryableError(
                "replace",
                document_id=document_id,
                attempt_token=attempt_token,
            ) from commit_error
        logger.warning(
            "metadata_commit_reconciliation outcome=unknown operation=%s "
            "kb_id=%s document_id=%s attempt_token=%s reason=%s "
            "commit_error_type=%s readback_error_type=%s",
            "document_replace_cow",
            kb_id,
            document_id,
            attempt_token,
            reconciliation.reason or "unknown",
            type(commit_error).__name__,
            type(readback_error).__name__ if readback_error is not None else None,
        )
        if caller_cancelled:
            if isinstance(commit_error, asyncio.CancelledError):
                raise commit_error
            raise asyncio.CancelledError() from commit_error
        raise DocumentCowCommitOutcomeUnknownError(
            "replace",
            document_id=document_id,
            job_id=job_id,
            attempt_token=attempt_token,
            manifest_group_id="",
            reason=reconciliation.reason,
        ) from (readback_error or commit_error)

    async def execute_document_delete_cow(
        self,
        kb_id: str,
        document_id: str,
        *,
        job_id: str,
        kb_generation: str,
        engine_delete: Callable[
            [str, str, str | None, str],
            Awaitable[dict[str, Any] | None],
        ],
        claim_token: str | None = None,
        expected_snapshot: str | None = None,
        delete_source_file: bool = True,
        delete_artifacts: bool = True,
        metadata_patch: dict[str, Any] | None = None,
    ) -> DocumentDeleteCowResult:
        """Execute one object-authoritative document deletion.

        The pre-engine recheck proves the delete has not already committed;
        the engine delete is idempotent; the post-engine commit is reconciled
        cancellation-safely.  No direct object/local byte cleanup is performed.
        """

        if not self.object_authoritative:
            raise DocumentCowError("Document delete COW requires object artifact mode")
        async with self.kb_write_guard(kb_id, expected_generation=kb_generation):
            record = await assert_active_kb_generation(
                self._kb_service,
                self._metadata_store,
                kb_id,
                kb_generation,
            )
            claim = await self._metadata_store.claim_document_deleting_cow(
                kb_id,
                document_id,
                kb_generation=kb_generation,
                job_id=job_id,
                claim_token=claim_token,
                expected_snapshot=expected_snapshot,
            )
            group_id, manifests = await self._prepare_document_cow_manifests(
                "delete",
                claim,
                job_id=job_id,
                kb_generation=kb_generation,
                retain_source=not delete_source_file,
                retain_artifacts=not delete_artifacts,
            )
            pre_engine = (
                await self._metadata_store.reconcile_document_delete_cow_commit(
                    kb_id,
                    document_id,
                    kb_generation=kb_generation,
                    job_id=job_id,
                    attempt_token=claim.attempt_token,
                    expected_snapshot=claim.snapshot_digest,
                    manifests=manifests,
                )
            )
            if pre_engine.outcome is MetadataCommitOutcome.COMMITTED:
                if pre_engine.value is not None:
                    return _document_delete_cow_result_from_commit(
                        pre_engine.value, claim.attempt_token
                    )
            elif pre_engine.outcome is MetadataCommitOutcome.UNKNOWN:
                raise DocumentCowCommitOutcomeUnknownError(
                    "delete",
                    document_id=document_id,
                    job_id=job_id,
                    attempt_token=claim.attempt_token,
                    manifest_group_id=group_id,
                    reason=pre_engine.reason,
                )
            engine_identity = _document_cow_engine_identity(
                kb_id=kb_id,
                document_id=document_id,
                source_generation_id=str(
                    claim.document.metadata.get("source_generation_id") or ""
                ),
            )
            engine_result = await self._run_document_cow_engine_delete(
                engine_delete,
                kb_id=kb_id,
                document_id=document_id,
                previous_lightrag_doc_id=claim.previous_lightrag_doc_id,
                engine_identity=engine_identity,
            )
            try:
                committed = await self._metadata_store.commit_document_delete_cow(
                    kb_id,
                    document_id,
                    kb_generation=kb_generation,
                    job_id=job_id,
                    attempt_token=claim.attempt_token,
                    expected_snapshot=claim.snapshot_digest,
                    metadata_patch={
                        **(
                            {"lightrag_delete_result": engine_result}
                            if engine_result is not None
                            else {}
                        ),
                        **(metadata_patch or {}),
                    },
                    manifests=manifests,
                )
            except (Exception, asyncio.CancelledError) as commit_error:
                return await self._reconcile_document_delete_cow_commit_exception(
                    kb_id=kb_id,
                    document_id=document_id,
                    kb_generation=kb_generation,
                    job_id=job_id,
                    attempt_token=claim.attempt_token,
                    expected_snapshot=claim.snapshot_digest,
                    manifests=manifests,
                    commit_error=commit_error,
                )
            _ = record
            return _document_delete_cow_result_from_commit(
                committed, claim.attempt_token
            )

    async def _reconcile_document_delete_cow_commit_exception(
        self,
        *,
        kb_id: str,
        document_id: str,
        kb_generation: str,
        job_id: str,
        attempt_token: str,
        expected_snapshot: str,
        manifests: tuple[ArtifactCleanupManifestRecord, ...],
        commit_error: BaseException,
    ) -> DocumentDeleteCowResult:
        caller_cancelled = isinstance(commit_error, asyncio.CancelledError)
        readback_error: BaseException | None = None
        try:
            safe = await await_cancellation_safe_reconciliation(
                lambda: self._metadata_store.reconcile_document_delete_cow_commit(
                    kb_id,
                    document_id,
                    kb_generation=kb_generation,
                    job_id=job_id,
                    attempt_token=attempt_token,
                    expected_snapshot=expected_snapshot,
                    manifests=manifests,
                )
            )
            reconciliation = safe.value
            caller_cancelled = caller_cancelled or safe.caller_cancelled
        except asyncio.CancelledError as exc:
            readback_error = exc
            reconciliation = MetadataCommitReconciliation[DocumentMutationCommitResult](
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="readback_cancelled",
            )
        except Exception as exc:  # noqa: BLE001
            readback_error = exc
            reconciliation = MetadataCommitReconciliation[DocumentMutationCommitResult](
                outcome=MetadataCommitOutcome.UNKNOWN,
                reason="readback_failed",
            )

        if (
            reconciliation.outcome is MetadataCommitOutcome.COMMITTED
            and reconciliation.value is not None
        ):
            return _document_delete_cow_result_from_commit(
                reconciliation.value, attempt_token
            )
        if reconciliation.outcome is MetadataCommitOutcome.ROLLED_BACK:
            # Bytes are preserved; the attempt remains fenced and retryable.
            if caller_cancelled:
                if isinstance(commit_error, asyncio.CancelledError):
                    raise commit_error
                raise asyncio.CancelledError() from commit_error
            raise DocumentCowRetryableError(
                "delete",
                document_id=document_id,
                attempt_token=attempt_token,
            ) from commit_error
        if caller_cancelled:
            if isinstance(commit_error, asyncio.CancelledError):
                raise commit_error
            raise asyncio.CancelledError() from commit_error
        raise DocumentCowCommitOutcomeUnknownError(
            "delete",
            document_id=document_id,
            job_id=job_id,
            attempt_token=attempt_token,
            manifest_group_id="",
            reason=reconciliation.reason,
        ) from (readback_error or commit_error)

    async def _run_document_cow_engine_delete(
        self,
        engine_delete: Callable[
            [str, str, str | None, str],
            Awaitable[dict[str, Any] | None],
        ],
        *,
        kb_id: str,
        document_id: str,
        previous_lightrag_doc_id: str | None,
        engine_identity: str,
    ) -> dict[str, Any] | None:
        if not previous_lightrag_doc_id:
            return None
        return await engine_delete(
            kb_id, document_id, previous_lightrag_doc_id, engine_identity
        )

    async def _document_cow_cleanup_counts(
        self,
        *,
        kb_id: str,
        manifest_group_id: str,
    ) -> tuple[int, int, int]:
        """Count pending/retained/blocked manifests for one resumed group.

        Used when re-driving a committed replace whose counts were not captured
        in the same process.  An absent group (empty manifest group) returns
        zeros.
        """

        if not manifest_group_id:
            return 0, 0, 0
        manifests, total = await self._metadata_store.list_artifact_cleanup_manifests(
            kb_id=kb_id,
            manifest_group_id=manifest_group_id,
            limit=ARTIFACT_LIFECYCLE_MAX_PAGE_SIZE,
        )
        if total > len(manifests):
            raise DocumentCowError(
                "Document COW manifest group exceeds the bounded page size"
            )
        pending = sum(1 for item in manifests if item.status in {"pending", "leased"})
        retained = sum(1 for item in manifests if item.status == "retained")
        blocked = sum(1 for item in manifests if item.status == "blocked")
        return pending, retained, blocked


def build_text_source(
    *, text: str, source_name: str | None = None, metadata: dict[str, Any] | None = None
) -> DocumentSourceInput:
    normalized_text = text.strip()
    if not normalized_text:
        raise ValueError("Text document cannot be empty")
    name = (
        source_name or f"text_{compute_mdhash_id(normalized_text, prefix='')[:12]}.txt"
    )
    return DocumentSourceInput(
        source_name=name,
        content=normalized_text.encode("utf-8"),
        source_type="text",
        content_type="text/plain; charset=utf-8",
        metadata=metadata or {},
    )


def _initial_source_document_commit_matches(
    candidate: DocumentRecord,
    persisted: DocumentRecord,
) -> bool:
    return bool(
        persisted.id == candidate.id
        and persisted.kb_id == candidate.kb_id
        and persisted.workspace == candidate.workspace
        and persisted.source_name == candidate.source_name
        and persisted.source_hash == candidate.source_hash
        and persisted.metadata.get("source_object_uri")
        == candidate.metadata.get("source_object_uri")
        and persisted.metadata.get("batch_id") == candidate.metadata.get("batch_id")
    )


def _initial_source_job_commit_matches(
    candidate: JobRecord,
    persisted: JobRecord,
) -> bool:
    return bool(
        persisted.id == candidate.id
        and persisted.kb_id == candidate.kb_id
        and persisted.workspace == candidate.workspace
        and persisted.batch_id == candidate.batch_id
        and persisted.document_id == candidate.document_id
        and persisted.job_type == candidate.job_type
        and persisted.idempotency_key == candidate.idempotency_key
        and persisted.total_items == candidate.total_items
        and persisted.payload.get("document_ids")
        == candidate.payload.get("document_ids")
        and persisted.payload.get("idempotency_fingerprint")
        == candidate.payload.get("idempotency_fingerprint")
    )


def _artifact_commit_candidate_matches(
    candidate: ArtifactRecord,
    persisted: ArtifactRecord,
) -> bool:
    return bool(
        persisted.id == candidate.id
        and persisted.kb_id == candidate.kb_id
        and persisted.workspace == candidate.workspace
        and persisted.document_id == candidate.document_id
        and persisted.artifact_type == candidate.artifact_type
        and persisted.uri == candidate.uri
        and persisted.checksum == candidate.checksum
        and persisted.metadata.get("object_uri") == candidate.metadata.get("object_uri")
        and persisted.metadata.get("object_prefix_uri")
        == candidate.metadata.get("object_prefix_uri")
    )


def _content_hash(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _idempotency_fingerprint(payload: dict[str, Any]) -> str:
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return f"sha256:{hashlib.sha256(encoded).hexdigest()}"


def _sanitize_source_name(source_name: str) -> str:
    clean_name = source_name.replace("..", "")
    clean_name = "".join(_sanitize_filename_char(char) for char in clean_name)
    clean_name = clean_name.strip().strip(".")
    if not clean_name:
        raise ValueError("Invalid document source name")
    return clean_name


def _write_source_file(
    workspace_dir: Path, document_id: str, filename: str, content: bytes
) -> Path:
    document_dir = workspace_dir / document_id
    document_dir.mkdir(mode=0o700, parents=False, exist_ok=False)
    target_path = document_dir / filename
    try:
        resolved_workspace = workspace_dir.resolve()
        resolved_target = target_path.resolve(strict=False)
        if not resolved_target.is_relative_to(resolved_workspace):
            raise ValueError("Document source path escapes workspace directory")
    except OSError as exc:
        raise ValueError("Invalid document source path") from exc

    with target_path.open("xb") as output:
        output.write(content)
        output.flush()
    return target_path


def _replacement_source_target(
    document_dir: Path, source_name: str, job_id: str
) -> Path:
    target_path = document_dir / source_name
    try:
        resolved_document_dir = document_dir.resolve(strict=False)
        resolved_target = target_path.resolve(strict=False)
        if not resolved_target.is_relative_to(resolved_document_dir):
            raise ValueError(
                "Document replacement source path escapes document directory"
            )
    except OSError as exc:
        raise ValueError("Invalid document replacement source path") from exc
    if not target_path.exists():
        return target_path
    return document_dir / f"{job_id}_{source_name}"


def _artifact_object_key(
    document: DocumentRecord, artifact: ArtifactRecord, path: Path
) -> str:
    return "/".join(
        [
            "workspaces",
            document.workspace,
            "documents",
            document.id,
            "artifacts",
            artifact.artifact_type,
            artifact.id,
            path.name,
        ]
    )


def _artifact_object_prefix(
    document: DocumentRecord, artifact: ArtifactRecord, path: Path
) -> str:
    return "/".join(
        [
            "workspaces",
            document.workspace,
            "documents",
            document.id,
            "artifacts",
            artifact.artifact_type,
            artifact.id,
            path.name,
        ]
    )


def _validate_document_cleanup_paths(
    workspace_dir: Path,
    document: DocumentRecord,
    artifacts: list[ArtifactRecord],
    *,
    delete_source_file: bool,
    delete_artifacts: bool,
) -> None:
    if not (delete_source_file or delete_artifacts):
        return
    source_path = _safe_workspace_path(workspace_dir, document.source_uri)
    document_dir = source_path.parent.resolve(strict=False)
    if delete_artifacts:
        for artifact in artifacts:
            _safe_document_path(workspace_dir, document_dir, artifact.uri)


def _resolve_artifact_path(
    source_root: Path, document: DocumentRecord, artifact: ArtifactRecord
) -> tuple[Path, bool]:
    """Return (path, is_directory) after running containment checks."""
    if not artifact.uri:
        raise ValueError("Artifact URI is empty")

    try:
        artifact_path = Path(artifact.uri).resolve(strict=False)
    except OSError as exc:
        raise ValueError("Invalid artifact path") from exc

    allowed_document_dir = (source_root / document.workspace / document.id).resolve(
        strict=False
    )
    if not artifact_path.is_relative_to(allowed_document_dir):
        raise ValueError("Artifact path escapes document directory")
    if artifact_path.exists():
        is_directory = artifact_path.is_dir()
        if not is_directory and not artifact_path.is_file():
            raise ValueError("Artifact is neither a file nor a directory")
        return artifact_path, is_directory
    object_uri = artifact.metadata.get("object_uri")
    object_prefix_uri = artifact.metadata.get("object_prefix_uri")
    if isinstance(object_uri, str) and object_uri:
        return artifact_path, False
    if isinstance(object_prefix_uri, str) and object_prefix_uri:
        return artifact_path, True
    raise FileNotFoundError(f"Artifact file not found: {artifact.id}")


def _artifact_logical_filename(
    document: DocumentRecord,
    artifact: ArtifactRecord,
    *,
    allow_uri: bool = True,
) -> str:
    candidates: list[object] = []
    if artifact.artifact_type == "original":
        candidates.extend([artifact.metadata.get("source_name"), document.source_name])
    candidates.extend(
        [artifact.metadata.get("filename"), artifact.metadata.get("source_name")]
    )
    if allow_uri and artifact.uri:
        parsed_path = (
            urlsplit(artifact.uri).path if "://" in artifact.uri else artifact.uri
        )
        candidates.append(Path(parsed_path.rstrip("/")).name)
    candidates.append(
        {
            "blocks": "blocks.jsonl",
            "preview_text": "preview.txt",
            "preview_table_json": "preview.table.json",
            "preview_html": "preview.html",
            "content_list": "content_list.json",
            "middle_json": "middle.json",
            "model_json": "model.json",
            "layout_pdf": "layout.pdf",
        }.get(artifact.artifact_type, artifact.artifact_type)
    )
    for candidate in candidates:
        if not isinstance(candidate, str) or not candidate.strip():
            continue
        try:
            return _sanitize_source_name(Path(candidate.strip()).name)
        except ValueError:
            continue
    return f"artifact-{artifact.id}"


def _artifact_materialization_target_name(
    artifact: ArtifactRecord, logical_filename: str, *, is_directory: bool
) -> str:
    digest = hashlib.sha256(artifact.id.encode("utf-8")).hexdigest()[:24]
    if is_directory:
        return f"artifact-{digest}"
    suffix = Path(logical_filename).suffix
    safe_suffix = "".join(
        char for char in suffix[:24] if char.isalnum() or char in {".", "_", "-"}
    )
    return f"artifact-{digest}{safe_suffix}"


def _safe_workspace_path(workspace_dir: Path, uri: str) -> Path:
    path = Path(uri)
    if not path.is_absolute():
        path = workspace_dir / path
    try:
        resolved = path.resolve(strict=False)
        resolved.relative_to(workspace_dir)
    except ValueError as exc:
        raise ValueError(f"Path is outside KB workspace: {uri}") from exc
    return resolved


def _safe_document_path(workspace_dir: Path, document_dir: Path, uri: str) -> Path:
    path = _safe_workspace_path(workspace_dir, uri)
    try:
        path.resolve(strict=False).relative_to(document_dir)
    except ValueError as exc:
        raise ValueError(f"Path is outside document directory: {uri}") from exc
    return path


def _remove_path(path: Path) -> None:
    if path.is_dir():
        shutil.rmtree(path)
    else:
        path.unlink(missing_ok=True)


def _remove_empty_parents(path: Path, *, stop_at: Path) -> None:
    current = path
    while current != stop_at and current.is_relative_to(stop_at):
        try:
            current.rmdir()
        except OSError:
            break
        current = current.parent


def _resolve_downloadable_artifact_path(
    source_root: Path, document: DocumentRecord, artifact: ArtifactRecord
) -> Path:
    if artifact.metadata.get("is_directory"):
        raise ValueError("Directory artifacts cannot be downloaded directly")
    path, is_directory = _resolve_artifact_path(source_root, document, artifact)
    if is_directory:
        raise ValueError("Artifact is not a downloadable file")
    return path


def _artifact_media_type(
    document: DocumentRecord,
    artifact: ArtifactRecord,
    path: Path,
    is_directory: bool = False,
) -> str:
    if is_directory:
        return "application/zip"
    metadata_media_type = artifact.metadata.get("media_type")
    if isinstance(metadata_media_type, str) and metadata_media_type:
        return metadata_media_type
    if artifact.artifact_type == "original" and document.content_type:
        return document.content_type
    if path.suffix.lower() == ".jsonl":
        return "application/x-ndjson"
    guessed_type, _encoding = mimetypes.guess_type(path.name)
    return guessed_type or "application/octet-stream"


def _artifact_route_url(
    kb_id: str, document_id: str, artifact_id: str, *, suffix: str
) -> str:
    return f"/kbs/{kb_id}/documents/{document_id}/artifacts/{artifact_id}{suffix}"


def _artifact_size_bytes(artifact: ArtifactRecord, path: Path) -> int | None:
    if artifact.size_bytes is not None:
        return artifact.size_bytes
    try:
        return path.stat().st_size if path.is_file() else None
    except OSError:
        return None


def _preview_kind_for_artifact_type(artifact_type: str) -> str:
    return {
        "preview_text": "text",
        "preview_table_json": "table",
        "preview_html": "html",
    }.get(artifact_type, artifact_type)


def _inline_preview_kind_for_media_type(media_type: str) -> str | None:
    normalized = media_type.split(";", 1)[0].lower()
    if normalized == "application/pdf":
        return "pdf"
    if normalized.startswith("image/") and normalized != "image/svg+xml":
        return "image"
    if normalized.startswith("text/"):
        return "text"
    if normalized in {
        "application/json",
        "application/ld+json",
        "application/markdown",
        "application/x-ndjson",
    }:
        return "text"
    return None


def _build_preview_artifacts(
    plan: DocumentParsePlan,
    execution: DocumentParseExecution,
    parsed_data: dict[str, Any],
    *,
    object_authoritative: bool,
) -> list[PendingArtifact]:
    source_path = execution.source_path
    preview_dir = execution.parsed_tree / f"{source_path.name}.preview"
    if preview_dir.exists():
        shutil.rmtree(preview_dir, ignore_errors=True)
    preview_dir.mkdir(parents=True, exist_ok=True)

    now = utc_now_iso()
    artifacts: list[PendingArtifact] = []

    table_artifact = _build_table_preview_artifact(
        plan,
        execution,
        source_path=source_path,
        preview_dir=preview_dir,
        created_at=now,
        object_authoritative=object_authoritative,
    )
    if table_artifact is not None:
        artifacts.append(table_artifact)

    text = parsed_data.get("content")
    if not isinstance(text, str) or not text.strip():
        text = _read_text_preview_source(source_path)
    if isinstance(text, str) and text.strip():
        preview_text, truncated = _truncate_utf8_text(text, _PREVIEW_TEXT_MAX_BYTES)
        text_path = preview_dir / f"{source_path.stem or 'document'}.preview.txt"
        text_path.write_text(preview_text, encoding="utf-8")
        artifacts.append(
            _pending_artifact(
                plan,
                execution,
                artifact_type="preview_text",
                runtime_path=text_path,
                created_at=now,
                object_authoritative=object_authoritative,
                metadata={
                    **_preview_metadata(plan, truncated=truncated),
                    "media_type": "text/plain; charset=utf-8",
                },
            )
        )

    if not artifacts:
        shutil.rmtree(preview_dir, ignore_errors=True)
    return artifacts


def _preview_metadata(plan: DocumentParsePlan, *, truncated: bool) -> dict[str, Any]:
    return {
        "preview": True,
        "source_hash": plan.document.source_hash,
        "parser_hash": plan.parser_hash,
        "parse_engine": plan.parser_engine,
        "truncated": truncated,
        "preview_schema_version": _PREVIEW_SCHEMA_VERSION,
    }


def _truncate_utf8_text(text: str, max_bytes: int) -> tuple[str, bool]:
    encoded = text.encode("utf-8")
    if len(encoded) <= max_bytes:
        return text, False
    return encoded[:max_bytes].decode("utf-8", errors="ignore"), True


def _read_text_preview_source(source_path: Path) -> str | None:
    if source_path.suffix.lower() not in _PREVIEW_TEXT_SUFFIXES:
        return None
    try:
        return source_path.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError):
        return None


def _build_table_preview_artifact(
    plan: DocumentParsePlan,
    execution: DocumentParseExecution,
    *,
    source_path: Path,
    preview_dir: Path,
    created_at: str,
    object_authoritative: bool,
) -> PendingArtifact | None:
    suffix = source_path.suffix.lower()
    payload: dict[str, Any] | None = None
    if suffix == ".csv":
        payload = _csv_preview_payload(source_path)
    elif suffix == ".xlsx":
        payload = _xlsx_preview_payload(source_path)
    if payload is None:
        return None

    table_path = preview_dir / f"{source_path.stem or 'document'}.preview.table.json"
    table_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return _pending_artifact(
        plan,
        execution,
        artifact_type="preview_table_json",
        runtime_path=table_path,
        created_at=created_at,
        object_authoritative=object_authoritative,
        metadata={
            **_preview_metadata(plan, truncated=bool(payload.get("truncated"))),
            "media_type": "application/json",
        },
    )


def _csv_preview_payload(source_path: Path) -> dict[str, Any] | None:
    try:
        with source_path.open("r", encoding="utf-8", newline="") as file:
            reader = csv.reader(file)
            rows: list[list[str]] = []
            truncated = False
            for row_index, row in enumerate(reader):
                if row_index >= _PREVIEW_TABLE_MAX_ROWS:
                    truncated = True
                    break
                if len(row) > _PREVIEW_TABLE_MAX_COLS:
                    truncated = True
                rows.append([str(value) for value in row[:_PREVIEW_TABLE_MAX_COLS]])
    except (OSError, UnicodeDecodeError, csv.Error):
        return None
    return _table_preview_payload(
        source_path.name,
        sheets=[{"name": source_path.stem or "Sheet1", "rows": rows}],
        truncated=truncated,
    )


def _xlsx_preview_payload(source_path: Path) -> dict[str, Any] | None:
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(source_path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - preview should not fail parse completion
        return None

    try:
        sheets: list[dict[str, Any]] = []
        truncated = False
        for sheet_index, sheet in enumerate(workbook.worksheets):
            if sheet_index >= 3:
                truncated = True
                break
            rows: list[list[str]] = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= _PREVIEW_TABLE_MAX_ROWS:
                    truncated = True
                    break
                values = ["" if value is None else str(value) for value in row]
                if len(values) > _PREVIEW_TABLE_MAX_COLS:
                    truncated = True
                rows.append(values[:_PREVIEW_TABLE_MAX_COLS])
            sheets.append({"name": str(sheet.title), "rows": rows})
        return _table_preview_payload(
            source_path.name, sheets=sheets, truncated=truncated
        )
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


def _table_preview_payload(
    source_name: str, *, sheets: list[dict[str, Any]], truncated: bool
) -> dict[str, Any]:
    return {
        "preview_schema_version": _PREVIEW_SCHEMA_VERSION,
        "kind": "table",
        "source_name": source_name,
        "truncated": truncated,
        "sheets": sheets,
    }


def _apply_parse_defaults(
    parser_engine: str | None,
    process_options: str | None,
    defaults: dict[str, str],
) -> tuple[str | None, str | None]:
    if parser_engine is None:
        parser_engine = defaults.get("parser_engine")
    if process_options is None:
        process_options = defaults.get("process_options")
    return parser_engine, process_options


def _first_present(*values: Any) -> Any:
    for value in values:
        if value is not None:
            return value
    return None


def _resolve_parse_directives(
    source_path: Path,
    document: DocumentRecord,
    *,
    parser_engine: str | None,
    process_options: str | None,
    active_parser_engine: str | None = None,
    active_process_options: str | None = None,
) -> tuple[str, str]:
    resolved_options: str | None = None
    if parser_engine is not None:
        engine = normalize_parser_engine(parser_engine)
    else:
        metadata_engine = document.metadata.get("parser_engine")
        if metadata_engine:
            engine = normalize_parser_engine(metadata_engine)
        elif active_parser_engine is not None:
            engine = normalize_parser_engine(active_parser_engine)
        else:
            engine, resolved_options = resolve_file_parser_directives(
                source_path, require_external_endpoint=False
            )

    if engine not in _PARSEABLE_ENGINES or engine not in SUPPORTED_PARSER_ENGINES:
        raise ValueError(f"Unsupported parser engine: {engine}")
    suffix = parser_suffix(source_path)
    if not parser_engine_supports_suffix(engine, suffix):
        raise ValueError(f"Parser engine '{engine}' does not support .{suffix} files")

    raw_options = _first_present(
        process_options,
        document.metadata.get("process_options"),
        active_process_options,
        resolved_options,
    )
    raw_options_text = "" if raw_options is None else str(raw_options)
    errors = validate_process_options(raw_options_text)
    if errors:
        raise ValueError("; ".join(errors))
    options = sanitize_process_options(raw_options_text)
    return engine, options


def _validate_parse_request_directives(
    *, parser_engine: str | None, process_options: str | None
) -> None:
    if parser_engine is not None:
        engine = normalize_parser_engine(parser_engine)
        if engine not in _PARSEABLE_ENGINES or engine not in SUPPORTED_PARSER_ENGINES:
            raise ValueError(f"Unsupported parser engine: {parser_engine}")
    raw_options_text = "" if process_options is None else str(process_options)
    errors = validate_process_options(raw_options_text)
    if errors:
        raise ValueError("; ".join(errors))


def _batch_parse_failure(
    document_id: str,
    *,
    error_code: str,
    error_message: str,
    existing_job_id: str | None = None,
) -> dict[str, Any]:
    failure: dict[str, Any] = {
        "document_id": document_id,
        "status": "failed",
        "error_code": error_code,
        "error_message": error_message,
    }
    if existing_job_id is not None:
        failure["existing_job_id"] = existing_job_id
    return failure


def _parser_hash(*, engine: str, process_options: str) -> str:
    payload = {
        "schema": "kb-parser-hash-v1",
        "engine": engine,
        "process_options": process_options,
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return f"sha256:{hashlib.sha256(encoded).hexdigest()}"


def _claimed_parse_artifact_binding(
    plan: DocumentParsePlan,
) -> PipelineArtifactBinding:
    if not plan.job_id:
        raise DocumentLifecycleError(
            "Parse artifact binding requires a claimed job identity"
        )
    if not plan.claim_token:
        raise DocumentLifecycleError(
            "Parse artifact binding requires a claimed attempt token"
        )
    if not plan.kb_generation:
        raise DocumentLifecycleError("Parse artifact binding requires a KB generation")
    return PipelineArtifactBinding(
        version=1,
        authority="kb_metadata",
        state="claimed",
        operation="parse",
        kb_id=plan.document.kb_id,
        kb_generation=plan.kb_generation,
        workspace=plan.document.workspace,
        document_id=plan.document.id,
        lightrag_doc_id=plan.lightrag_doc_id,
        job_id=plan.job_id,
        claim_token=plan.claim_token,
        source_hash=plan.document.source_hash or None,
        parser_hash=plan.parser_hash,
        parse_generation_id=plan.claim_token,
        index_hash=plan.document.index_hash,
        sidecar_artifact_id=plan.expected_current_sidecar_artifact_id,
        blocks_artifact_id=plan.expected_current_blocks_artifact_id,
        expected_current_sidecar_artifact_id=(
            plan.expected_current_sidecar_artifact_id
        ),
        expected_current_blocks_artifact_id=(plan.expected_current_blocks_artifact_id),
        raw_artifact_ids=tuple(ref.artifact_id for ref in plan.raw_object_refs),
    )


def _validate_parse_artifact_binding(
    binding: PipelineArtifactBinding,
    plan: DocumentParsePlan,
    *,
    require_claimed: bool,
) -> None:
    expected_identities: dict[str, object] = {
        "operation": "parse",
        "kb_id": plan.document.kb_id,
        "kb_generation": plan.kb_generation,
        "workspace": plan.document.workspace,
        "document_id": plan.document.id,
        "lightrag_doc_id": plan.lightrag_doc_id,
        "job_id": plan.job_id,
        "claim_token": plan.claim_token,
        "source_hash": plan.document.source_hash or None,
        "parser_hash": plan.parser_hash,
        "expected_current_sidecar_artifact_id": (
            plan.expected_current_sidecar_artifact_id
        ),
        "expected_current_blocks_artifact_id": (
            plan.expected_current_blocks_artifact_id
        ),
    }
    mismatches = [
        field_name
        for field_name, expected in expected_identities.items()
        if getattr(binding, field_name) != expected
    ]
    expected_raw_ids = tuple(ref.artifact_id for ref in plan.raw_object_refs)
    if binding.raw_artifact_ids != expected_raw_ids:
        mismatches.append("raw_artifact_ids")
    if require_claimed:
        if binding.state != "claimed":
            mismatches.append("state")
        if binding.parse_generation_id != plan.claim_token:
            mismatches.append("parse_generation_id")
        if binding.sidecar_artifact_id != plan.expected_current_sidecar_artifact_id:
            mismatches.append("sidecar_artifact_id")
        if binding.blocks_artifact_id != plan.expected_current_blocks_artifact_id:
            mismatches.append("blocks_artifact_id")
    if mismatches:
        raise DocumentLifecycleError(
            "Parse artifact binding does not match the claimed attempt: "
            + ", ".join(sorted(set(mismatches)))
        )


def _durable_binding_full_doc_payload(
    existing: dict[str, Any],
    *,
    binding: PipelineArtifactBinding,
    file_path: str,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "content": existing.get("content", ""),
        "file_path": canonicalize_pipeline_logical_filename(file_path),
        "parse_format": existing.get("parse_format", FULL_DOCS_FORMAT_LIGHTRAG),
        "artifact_binding": binding.to_dict(),
    }
    for key in (
        "parse_engine",
        "process_options",
        "chunk_options",
        "content_hash",
        "update_time",
    ):
        value = existing.get(key)
        if value is not None:
            payload[key] = value
    return payload


def _build_parse_artifacts(
    plan: DocumentParsePlan,
    execution: DocumentParseExecution,
    parsed_data: dict[str, Any],
    *,
    object_authoritative: bool,
) -> list[PendingArtifact]:
    now = utc_now_iso()
    artifacts = [
        _pending_artifact(
            plan,
            execution,
            artifact_type="original",
            runtime_path=execution.source_path,
            created_at=now,
            object_authoritative=object_authoritative,
            metadata={
                "source_name": plan.document.source_name,
                "content_type": plan.document.content_type,
                "source_hash": plan.document.source_hash,
            },
        )
    ]

    blocks_path_value = parsed_data.get("blocks_path")
    blocks_path = Path(blocks_path_value) if blocks_path_value else None
    sidecar_dir = blocks_path.parent if blocks_path is not None else None
    if sidecar_dir is not None and sidecar_dir.exists():
        durable_blocks_path = _durable_blocks_path(
            execution,
            blocks_path,
            object_authoritative=object_authoritative,
        )
        artifacts.append(
            _pending_artifact(
                plan,
                execution,
                artifact_type="sidecar",
                runtime_path=sidecar_dir,
                created_at=now,
                object_authoritative=object_authoritative,
                metadata={
                    "is_directory": True,
                    "blocks_path": durable_blocks_path,
                    "parse_engine": plan.parser_engine,
                    "parser_hash": plan.parser_hash,
                },
            )
        )
    if blocks_path is not None and blocks_path.exists():
        artifacts.append(
            _pending_artifact(
                plan,
                execution,
                artifact_type="blocks",
                runtime_path=blocks_path,
                created_at=now,
                object_authoritative=object_authoritative,
                metadata={"parse_engine": plan.parser_engine},
            )
        )

    if sidecar_dir is not None:
        for raw_dir in _raw_artifact_dirs(sidecar_dir, plan.parser_engine):
            if not raw_dir.exists():
                continue
            artifacts.append(
                _pending_artifact(
                    plan,
                    execution,
                    artifact_type="raw_dir",
                    runtime_path=raw_dir,
                    created_at=now,
                    object_authoritative=object_authoritative,
                    metadata={
                        "is_directory": True,
                        "parse_engine": plan.parser_engine,
                        **(
                            {"raw_directory_name": raw_dir.name}
                            if object_authoritative
                            else {}
                        ),
                    },
                )
            )
            artifacts.extend(
                _fine_grained_artifacts(
                    plan,
                    execution,
                    root=raw_dir,
                    created_at=now,
                    source="raw_dir",
                    object_authoritative=object_authoritative,
                )
            )
        artifacts.extend(
            _fine_grained_artifacts(
                plan,
                execution,
                root=sidecar_dir,
                created_at=now,
                source="sidecar",
                object_authoritative=object_authoritative,
            )
        )
    return artifacts


def _fine_grained_artifacts(
    plan: DocumentParsePlan,
    execution: DocumentParseExecution,
    *,
    root: Path,
    created_at: str,
    source: str,
    object_authoritative: bool,
) -> list[PendingArtifact]:
    if not root.is_dir():
        return []

    artifacts: list[PendingArtifact] = []
    seen: set[Path] = set()
    root_resolved = root.resolve(strict=False)

    def add(artifact_type: str, path: Path) -> None:
        if path.is_symlink() or not path.is_file():
            return
        resolved = path.resolve(strict=False)
        try:
            resolved.relative_to(root_resolved)
        except ValueError:
            return
        if resolved in seen:
            return
        seen.add(resolved)
        artifacts.append(
            _pending_artifact(
                plan,
                execution,
                artifact_type=artifact_type,
                runtime_path=path,
                created_at=created_at,
                object_authoritative=object_authoritative,
                metadata={
                    "parse_engine": plan.parser_engine,
                    "parser_hash": plan.parser_hash,
                    "source": source,
                    "relative_path": path.relative_to(root).as_posix(),
                },
            )
        )

    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        artifact_type = _fine_grained_artifact_type(path)
        if artifact_type is not None:
            add(artifact_type, path)

    return artifacts


def _fine_grained_artifact_type(path: Path) -> str | None:
    name = path.name.lower()
    suffix = path.suffix.lower()
    if name == _RAW_BUNDLE_MANIFEST:
        return None
    if name in _ROOT_FILE_ARTIFACT_TYPES:
        return _ROOT_FILE_ARTIFACT_TYPES[name]
    if name.endswith("_content_list.json"):
        return "content_list"
    if name.endswith("_middle.json") or name.endswith("_middle_json.json"):
        return "middle_json"
    if name.endswith("_model.json") or name.endswith("_model_json.json"):
        return "model_json"
    if suffix in _MARKDOWN_SUFFIXES:
        return "markdown"
    if suffix in _IMAGE_SUFFIXES:
        return "image"
    return None


def _pending_artifact(
    plan: DocumentParsePlan,
    execution: DocumentParseExecution,
    *,
    artifact_type: str,
    runtime_path: Path,
    created_at: str,
    object_authoritative: bool,
    metadata: dict[str, Any],
) -> PendingArtifact:
    is_file = runtime_path.is_file()
    is_directory = runtime_path.is_dir()
    durable_path = (
        execution.canonical_path_for(runtime_path)
        if object_authoritative
        else runtime_path
    )
    durable_metadata = dict(metadata)
    if object_authoritative:
        durable_metadata.setdefault("filename", runtime_path.name)
    return PendingArtifact(
        record=ArtifactRecord(
            id=generate_track_id(f"artifact_{artifact_type}"),
            kb_id=plan.document.kb_id,
            workspace=plan.document.workspace,
            document_id=plan.document.id,
            artifact_type=artifact_type,
            uri=str(durable_path),
            checksum=(
                _file_checksum(runtime_path)
                if is_file
                else _directory_checksum(runtime_path)
                if is_directory and object_authoritative
                else None
            ),
            size_bytes=runtime_path.stat().st_size if is_file else None,
            metadata=durable_metadata,
            created_at=created_at,
        ),
        runtime_path=runtime_path,
        is_directory=is_directory,
    )


def _raw_artifact_dirs(sidecar_dir: Path, engine: str) -> list[Path]:
    if not sidecar_dir.name.endswith(PARSED_DIR_SUFFIX):
        return []
    base = sidecar_dir.name[: -len(PARSED_DIR_SUFFIX)]
    if engine == PARSER_ENGINE_MINERU:
        return [sidecar_dir.parent / f"{base}{MINERU_RAW_DIR_SUFFIX}"]
    if engine == PARSER_ENGINE_DOCLING:
        return [
            sidecar_dir.parent / f"{base}{DOCLING_RAW_DIR_SUFFIX}",
            sidecar_dir.parent / f"{base}{LIBREOFFICE_RAW_DIR_SUFFIX}",
        ]
    return []


def _raw_directory_names(source_name: str, engine: str) -> list[str]:
    canonical_name = canonicalize_parser_hinted_basename(source_name) or source_name
    if engine == PARSER_ENGINE_MINERU:
        return [f"{canonical_name}{MINERU_RAW_DIR_SUFFIX}"]
    if engine == PARSER_ENGINE_DOCLING:
        return [
            f"{canonical_name}{DOCLING_RAW_DIR_SUFFIX}",
            f"{canonical_name}{LIBREOFFICE_RAW_DIR_SUFFIX}",
        ]
    return []


def _durable_blocks_path(
    execution: DocumentParseExecution,
    value: object,
    *,
    object_authoritative: bool,
) -> str | None:
    if value is None or value == "":
        return None
    runtime_path = Path(str(value))
    if object_authoritative:
        return str(execution.canonical_path_for(runtime_path))
    return str(runtime_path)


def _normalized_sha256_checksum(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = value.strip()
    if not normalized:
        return None
    digest = (
        normalized[len("sha256:") :] if normalized.startswith("sha256:") else normalized
    )
    if len(digest) != 64 or any(
        character not in "0123456789abcdefABCDEF" for character in digest
    ):
        return None
    return f"sha256:{digest.lower()}"


def _verify_document_source_checksum(document: DocumentRecord, path: Path) -> None:
    raw_expected = document.source_hash
    if raw_expected is None or not str(raw_expected).strip():
        # Historical rows can predate persisted source hashes.
        return
    expected = _normalized_sha256_checksum(raw_expected)
    if expected is None:
        raise DocumentSourceChecksumError(
            f"Document source checksum metadata is invalid for document '{document.id}'"
        )
    try:
        actual = _file_checksum(path)
    except OSError as exc:
        raise DocumentSourceChecksumError(
            f"Document source checksum could not be verified for document "
            f"'{document.id}'"
        ) from exc
    if actual != expected:
        raise DocumentSourceChecksumError(
            f"Document source checksum mismatch for document '{document.id}': "
            f"expected {expected}, got {actual}"
        )


def _materialized_raw_cache_matches(path: Path, expected_checksum: object) -> bool:
    expected = _normalized_sha256_checksum(expected_checksum)
    if expected is None:
        return False
    try:
        return _directory_checksum(path) == expected
    except (DocumentLifecycleError, OSError):
        return False


def _remove_materialized_raw_cache(path: Path, *, document_root: Path) -> None:
    root = document_root.resolve(strict=True)
    if path.is_symlink():
        path.unlink(missing_ok=True)
        return
    if not path.exists():
        return
    resolved = path.resolve(strict=True)
    if resolved == root or not resolved.is_relative_to(root):
        raise DocumentLifecycleError(
            "Materialized raw cache escapes its document lease"
        )
    if resolved.is_dir():
        shutil.rmtree(resolved)
    else:
        resolved.unlink(missing_ok=True)


def _file_checksum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return f"sha256:{digest.hexdigest()}"


def _directory_checksum(path: Path) -> str:
    digest = hashlib.sha256()
    root = path.resolve(strict=True)
    for child in sorted(root.rglob("*"), key=lambda item: item.as_posix()):
        if child.is_symlink() or not child.is_file():
            continue
        resolved = child.resolve(strict=True)
        if not resolved.is_relative_to(root):
            raise DocumentLifecycleError("Artifact directory entry escapes its root")
        relative = resolved.relative_to(root).as_posix().encode("utf-8")
        digest.update(len(relative).to_bytes(4, "big"))
        digest.update(relative)
        with resolved.open("rb") as file:
            for chunk in iter(lambda: file.read(1024 * 1024), b""):
                digest.update(chunk)
    return f"sha256:{digest.hexdigest()}"


def normalize_artifact_checksum(value: object) -> str | None:
    """Public checksum normalizer shared by processing-owner materialization."""

    return _normalized_sha256_checksum(value)


def file_artifact_checksum(path: Path) -> str:
    """Return the canonical checksum used by persisted file artifact rows."""

    return _file_checksum(path)


def directory_artifact_checksum(path: Path) -> str:
    """Return the canonical checksum used by persisted directory artifacts."""

    return _directory_checksum(path)
